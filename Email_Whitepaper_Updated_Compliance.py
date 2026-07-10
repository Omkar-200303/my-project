import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import re
import threading
from datetime import datetime
from playwright.sync_api import sync_playwright
import win32com.client
import openpyxl
import getpass
import socket
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─────────────────────────────────────────────
# FIND SYSTEM CHROME / EDGE
# ─────────────────────────────────────────────
def find_chrome_path():
    possible_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.join(os.environ.get("LOCALAPPDATA", ""),
                     r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES", ""),
                     r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", ""),
                     r"Google\Chrome\Application\chrome.exe"),
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        os.path.join(os.environ.get("PROGRAMFILES", ""),
                     r"Microsoft\Edge\Application\msedge.exe"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", ""),
                     r"Microsoft\Edge\Application\msedge.exe"),
    ]
    for p in possible_paths:
        if p and os.path.exists(p):
            return p
    return None

# ─────────────────────────────────────────────
# THEME
# ─────────────────────────────────────────────
class Theme:
    BG_BASE          = "#F7F4EF"
    BG_WHITE         = "#FFFFFF"
    BG_CARD          = "#FFFFFF"
    BG_INPUT         = "#F2EFE9"
    BG_HOVER         = "#F0EDE7"
    BG_ACCENT        = "#6C5CE7"
    BG_ACCENT_LIGHT  = "#EDE9FE"
    BG_SUCCESS       = "#10B981"
    BG_SUCCESS_LIGHT = "#D1FAE5"
    BG_DANGER        = "#EF4444"
    BG_DANGER_LIGHT  = "#FEE2E2"
    BG_WARNING       = "#F59E0B"
    BG_WARNING_LIGHT = "#FEF3C7"
    BG_INFO          = "#3B82F6"
    BG_INFO_LIGHT    = "#DBEAFE"
    BG_BLUE_BTN      = "#4F46E5"
    BG_GREEN_BTN     = "#059669"
    BG_ORANGE_BTN    = "#D97706"
    BG_TEAL_BTN      = "#0D9488"
    BG_CONSOLE       = "#1E1E2E"

    FG_PRIMARY   = "#1F2937"
    FG_SECONDARY = "#6B7280"
    FG_MUTED     = "#9CA3AF"
    FG_ACCENT    = "#6C5CE7"
    FG_WHITE     = "#FFFFFF"
    FG_INPUT     = "#374151"
    FG_SUCCESS   = "#059669"
    FG_DANGER    = "#DC2626"
    FG_WARNING   = "#D97706"

    BORDER       = "#E5E1DA"
    BORDER_LIGHT = "#F0EDE7"
    BORDER_FOCUS = "#6C5CE7"
    BORDER_CARD  = "#E8E4DD"
    SHADOW       = "#D6D2CB"

    FONT_FAMILY   = "Segoe UI"
    FONT_TITLE    = (FONT_FAMILY, 17, "bold")
    FONT_SUBTITLE = (FONT_FAMILY, 11, "bold")
    FONT_HEADING  = (FONT_FAMILY, 10, "bold")
    FONT_BODY     = (FONT_FAMILY, 9)
    FONT_SMALL    = (FONT_FAMILY, 8)
    FONT_TINY     = (FONT_FAMILY, 7)
    FONT_MONO     = ("Consolas", 9)
    FONT_ICON     = (FONT_FAMILY, 13)

# ─────────────────────────────────────────────
# AUTO-MAPPING
# ─────────────────────────────────────────────
AUTO_MAP = {
    "Asset": [
        "asset_downloaded__c","asset_downloaded_c","asset_downloaded",
        "asset_title__c","asset_title_c","asset_title_2","asset_title2",
        "asset_title","asset_name","asset_name__c",
        "second_touch_asset__c","second_touch_asset_c","second_touch_asset",
        "asset","asset name","asset 1","asset1",
        "content asset","content_asset","content name","content_name",
        "gated content name","gated_content_name","gated content","gated_content",
        "whitepaper","white paper","white_paper",
        "ebook","e-book","e_book",
        "resource","resource name","resource_name",
        "document","document name","document_name",
        "offer","offer name","offer_name",
        "assetdownloaded","asset download","asset_download",
        "downloaded asset","downloaded_asset",
        "content downloaded","content_downloaded",
        "report","report name","report_name",
        "guide","guide name","guide_name",
        "webinar","webinar name","webinar_name",
        "campaign asset","campaign_asset",
    ],
    "Email": [
        "email","email address","email_address",
        "actual email","actual_email",
        "work email","work_email",
        "business email","business_email",
        "company email","company_email",
        "contact email","contact_email",
        "primary email","primary_email",
        "corporate email","corporate_email",
        "emailaddress","email addr","email_addr",
        "e-mail","e_mail","e mail",
        "workemail","mail","mail address","mail_address",
        "email*","email id","email_id",
        "respondent email","respondent_email",
        "lead email","lead_email",
        "prospect email","prospect_email",
        "subscriber email","subscriber_email",
        "user email","user_email",
    ],
    "Country": [
        "country","country name","country_name",
        "country*","countries",
        "country_lead","lead country","lead_country",
        "companycountry","company country","company_country",
        "billing country","billing_country",
        "mailing country","mailing_country",
        "country/region","country_region","country region",
        "countryname","nation","nationality",
        "geo","geography","location country","location_country",
        "respondent country","respondent_country",
        "hq country","hq_country",
        "country of residence","country_of_residence",
        "operating country","operating_country",
    ],
    "Campaign": [
        "campaign","campaign name","campaign_name",
        "campaignname","campaign title","campaign_title",
        "utm_campaign","utm campaign",
        "campaign_id","campaign id",
        "campaign*","marketing campaign","marketing_campaign",
        "program","program name","program_name",
        "initiative","initiative name","initiative_name",
        "promotion","promotion name","promotion_name",
        "campaign description","campaign_description",
        "campaign source","campaign_source",
        "source campaign","source_campaign",
    ],
    "OLI/CID": [
        "oli","cid","oli/cid","oli_cid",
        "oli id","oli_id","cid id","cid_id",
        "order line item","order_line_item",
        "orderlineitem","order line","order_line",
        "campaign id","campaign_id",
        "line item","line_item","lineitem",
        "placement id","placement_id",
        "ad id","ad_id",
        "offer id","offer_id",
        "reference id","reference_id",
        "ref id","ref_id",
        "tracking id","tracking_id",
        "deal id","deal_id",
        "package id","package_id",
    ],
    "State/Region (optional)": [
        "state","state name","state_name",
        "state/province","state_province",
        "state/county","state_county",
        "state or province","stateprovince",
        "state/region","state_region",
        "region","region name","region_name",
        "province","province name","province_name",
        "territory","territory name","territory_name",
        "county","county name","county_name",
        "district","district name","district_name",
        "billing state","billing_state",
        "mailing state","mailing_state",
        "respondent state","respondent_state",
        "location state","location_state",
        "geo region","geo_region",
    ],
    "Segment": [
        "campaignname","campaign_name","campaign name",
        "segment","segment name","segment_name",
        "segmentname","seg","seg name","seg_name",
        "audience segment","audience_segment",
        "market segment","market_segment",
        "customer segment","customer_segment",
        "vertical","vertical name","vertical_name",
        "industry segment","industry_segment",
        "account segment","account_segment",
        "tier","tier name","tier_name",
        "segmentation","segmentation_name",
    ],
}

# ─────────────────────────────────────────────
# GLOBALS
# ─────────────────────────────────────────────
extracted_metadata = {"campaign": "Unknown", "oli": ""}

EUROPE_COUNTRIES = {
    'france','italy','spain','netherlands','belgium',
    'portugal','ireland','luxembourg','monaco','andorra',
    'liechtenstein','malta','san marino','vatican city',
    'sweden','norway','denmark','finland','iceland',
    'estonia','latvia','lithuania',
    'poland','czech republic','slovakia','hungary','romania',
    'bulgaria','moldova','ukraine','belarus','russia',
    'slovenia','croatia','bosnia and herzegovina','serbia',
    'montenegro','north macedonia','albania','kosovo',
    'greece','cyprus',
    'turkey','georgia','armenia','azerbaijan','kazakhstan',
    'uk','united kingdom','great britain','england',
    'scotland','wales','northern ireland',
    'czechia','russian federation',
    'bih','bosnia','fyrom','macedonia',
    'republic of moldova','republic of ireland',
    'republic of cyprus','republic of malta',
    'republic of serbia','republic of albania',
    'republic of kosovo','republic of montenegro',
    'republic of slovenia','republic of croatia',
    'republic of bulgaria','republic of hungary',
    'republic of poland','republic of slovakia',
    'republic of estonia','republic of latvia',
    'republic of lithuania','republic of finland',
    'kingdom of norway','kingdom of sweden',
    'kingdom of denmark','kingdom of spain',
    'kingdom of belgium','kingdom of netherlands',
    'grand duchy of luxembourg','principality of monaco',
    'principality of andorra','principality of liechtenstein',
    'republic of iceland','republic of portugal',
    'italian republic','french republic','hellenic republic',
    'republic of turkey','republic of georgia',
    'republic of armenia','republic of azerbaijan',
    'republic of kazakhstan','republic of belarus',
    'republic of ukraine',
    'fr','it','es','nl','be','pt','ie','lu','mc','ad',
    'li','mt','sm','va',
    'se','no','dk','fi','is','ee','lv','lt',
    'pl','cz','sk','hu','ro','bg','md','ua','by','ru',
    'si','hr','ba','rs','me','mk','al','xk',
    'gr','cy',
    'tr','ge','am','az','kz',
    'gb','uk',
}

DO_COUNTRIES = {
    'germany','switzerland','austria',
    'de','ch','at',
    'federal republic of germany',
    'swiss confederation',
    'republic of austria',
    'deutschland','schweiz','österreich','oesterreich',
}

ALL_EURO_DO = EUROPE_COUNTRIES | DO_COUNTRIES

CANADA_VARIANTS = {
    'canada','can',
    'dominion of canada',
}

US_VARIANTS = {
    'usa','us','united states','united states of america',
    'u.s.','u.s.a.','america',
}

GARTNER_KEYWORDS = {'gartner','gartner inc','gartner inc.'}

PORTAL_CONFIG = {
    "login_url"  : "https://admin.clickdotemail.com/",
    "lists_url"  : "https://admin.clickdotemail.com/EmailMarketing/views/mylists",
    "email_sel"  : ['input[type="email"]','input[name="email"]',
                    'input[id="email"]','input[placeholder*="email" i]',
                    'input[autocomplete="email"]','#email',
                    'input[type="text"]'],
    "pass_sel"   : ['input[type="password"]','input[name="password"]',
                    'input[id="password"]','input[placeholder*="password" i]',
                    '#password'],
    "signin_sel" : ['button:has-text("Sign In")','button:has-text("Login")',
                    'button:has-text("Log In")','button[type="submit"]',
                    'input[type="submit"]','#loginBtn'],
}

WAIT_AFTER_LOGIN        = 2
WAIT_AFTER_NAV          = 1
WAIT_AFTER_CREATE_CLICK = 1
WAIT_AFTER_SINGLE_LIST  = 1.5
WAIT_AFTER_CREATE_BTN   = 1.5
WAIT_AFTER_FILE_ATTACH  = 1.5
WAIT_AFTER_IMPORT       = 3

# ─────────────────────────────────────────────
# STOP WORDS
# ─────────────────────────────────────────────
STOP_WORDS = {
    'a','an','the','in','on','at','to','for','of','with','by','from',
    'into','through','about','between','during','before','after','above',
    'below','up','down','out','off','over','under','again','further',
    'then','once','and','but','or','nor','so','yet','both','either',
    'neither','not','only','own','same','than','too','very','just',
    'because','as','until','while','i','me','my','we','our','you',
    'your','he','she','it','its','they','them','their','this','that',
    'these','those','what','which','who','whom','is','are','was','were',
    'be','been','being','have','has','had','do','does','did','will',
    'would','shall','should','may','might','must','can','could','get',
    'got','more','most','other','some','such','no','nor','any','each',
    'few','how','all','also','here','there','when','where','why','now',
    'new','vs','via','per','etc','eg','ie',
}

# ─────────────────────────────────────────────
# SYSTEM USER
# ─────────────────────────────────────────────
def get_system_user_info():
    info = {"employee_id": "", "full_name": "", "domain": "", "pc_name": ""}
    try:
        info["pc_name"] = socket.gethostname()
    except:
        info["pc_name"] = "Unknown-PC"
    try:
        info["employee_id"] = os.getlogin()
    except:
        pass
    if not info["employee_id"]:
        try:
            info["employee_id"] = getpass.getuser()
        except:
            pass
    if not info["employee_id"]:
        info["employee_id"] = (os.environ.get("USERNAME") or
                               os.environ.get("USER") or "Unknown")
    info["domain"] = os.environ.get("USERDOMAIN", "")
    try:
        r = subprocess.run(["net", "user", info["employee_id"]],
                           capture_output=True, text=True, timeout=5)
        for line in r.stdout.splitlines():
            if "full name" in line.lower():
                parts = line.split(None, 2)
                info["full_name"] = (parts[2].strip() if len(parts) >= 3 else
                                     parts[1].strip() if len(parts) == 2 else "")
                break
    except:
        pass
    if not info["full_name"]:
        info["full_name"] = info["employee_id"]
    return info

# ─────────────────────────────────────────────
# CORE REGEX PATTERNS
# ─────────────────────────────────────────────
_clean_re        = re.compile(r'[^\w\s]')
_spaces_re       = re.compile(r'[\s]+')
_safe_re         = re.compile(r'[\\/*?:"\'<>|]')
_nonprint_re     = re.compile(r'[^\x20-\x7E]')
_illegal_path_re = re.compile(r'[\\/*?:"<>|]')

# Characters that break portal input fields — & , ; ! @ # etc.
_PORTAL_ILLEGAL_RE = re.compile(r'[&,;!@#$%^*()+={}\[\]|\\<>?/~`"\']')
_MULTI_UNDERSCORE  = re.compile(r'_+')

# ─────────────────────────────────────────────
# SANITIZE FOR PORTAL INPUT  ← KEY FIX
# ─────────────────────────────────────────────
def sanitize_for_portal(name: str, max_len: int = 150) -> str:
    """
    Remove / replace ALL characters that break the portal list-name
    input field.

    Removes  : & , ; ! @ # $ % ^ * ( ) + = { } [ ] | \\ < > ? / ~ ` " '
    Keeps    : letters, digits, underscore, hyphen, dot
    Collapses: multiple underscores → single underscore
    Truncates: to max_len characters
    """
    if not name:
        return "WP_Unknown"
    cleaned = _PORTAL_ILLEGAL_RE.sub('_', str(name))
    cleaned = _nonprint_re.sub('', cleaned)
    cleaned = _MULTI_UNDERSCORE.sub('_', cleaned)
    cleaned = cleaned.strip('_. ')
    cleaned = cleaned[:max_len].strip('_. ')
    return cleaned if cleaned else "WP_Unknown"

# ─────────────────────────────────────────────
# OLI NORMALIZER
# ─────────────────────────────────────────────
def normalize_oli(value) -> str:
    """Return '' for any non-meaningful OLI. Never returns '0000'."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan","none","0000","n/a","na","null","","0","unknown"):
        return ""
    if re.fullmatch(r'0+', s):
        return ""
    return s

# ─────────────────────────────────────────────
# CAMPAIGN VALIDATOR
# ─────────────────────────────────────────────
_INVALID_CAMPAIGNS = {
    "unknown","unknown_campaign","unknown campaign",
    "none","nan","null","","n/a","na","campaign",
}

def is_valid_campaign(value: str) -> bool:
    if not value:
        return False
    return value.strip().lower() not in _INVALID_CAMPAIGNS

# ─────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────
def clean_text(t):
    return _safe_re.sub('', str(t)).strip().replace(" ", "_")

def get_first_word(t):
    c = _clean_re.sub('', str(t)).strip()
    return c.split()[0] if c else "Unknown"

def get_asset_name_no_stopwords(text, max_words=5):
    """
    Build asset name for filename/portal.
    Strips ALL special characters including & , ; etc.
    """
    # Strip portal-illegal chars first
    cleaned = _PORTAL_ILLEGAL_RE.sub(' ', str(text)).strip()
    cleaned = _safe_re.sub(' ', cleaned).strip()
    cleaned = _nonprint_re.sub('', cleaned)

    raw_words = [w for w in _spaces_re.split(cleaned) if w]
    filtered  = [w for w in raw_words if w.lower() not in STOP_WORDS]
    selected  = filtered[:max_words] if filtered else raw_words[:max_words]
    result    = ''.join(selected)
    result    = _MULTI_UNDERSCORE.sub('_', result).strip('_')
    result    = _nonprint_re.sub('', result)
    return result if result else "Unknown_Asset"

def safe_folder_name(text, max_len=60):
    if not text:
        return "Unknown"
    cleaned = _illegal_path_re.sub('_', str(text))
    cleaned = _PORTAL_ILLEGAL_RE.sub('_', cleaned)
    cleaned = _nonprint_re.sub('', cleaned)
    cleaned = re.sub(r'[\s]+', '_', cleaned)
    cleaned = _MULTI_UNDERSCORE.sub('_', cleaned)
    cleaned = cleaned.strip('. _')
    cleaned = cleaned[:max_len]
    cleaned = cleaned.strip('. _')
    return cleaned if cleaned else "Unknown"

def safe_filename(text, max_len=50):
    cleaned = _illegal_path_re.sub('_', str(text))
    cleaned = _PORTAL_ILLEGAL_RE.sub('_', cleaned)
    cleaned = _nonprint_re.sub('', cleaned)
    cleaned = re.sub(r'[\s]+', '_', cleaned)
    cleaned = _MULTI_UNDERSCORE.sub('_', cleaned)
    cleaned = cleaned.strip('. _')[:max_len]
    cleaned = cleaned.strip('. _')
    return cleaned or "Email_Attachment"

# ─────────────────────────────────────────────
# EMAIL BODY PARSER — LABEL PRIORITY LISTS
# ─────────────────────────────────────────────
_CAMPAIGN_LABEL_PRIORITY = [
    "Campaign Name","Campaign name","campaign name",
    "CampaignName","campaignname","Campaign",
    # Fallback → Project Name used AS campaign
    "Project Name","Project name","project name",
    "ProjectName","projectname","Project",
]

_OLI_LABEL_PRIORITY = [
    "OLI/CID","OLI / CID","oli/cid",
    "OLI","oli","CID","cid",
    "Order Line Item","order line item",
]

_CLIENT_LABEL_PRIORITY = [
    "Client Name","Client name","client name","ClientName",
    "Client","client",
    "Company Name","Company name","company name",
    "Company","company",
    "Account Name","account name",
    "Account","account",
    "Advertiser","advertiser",
    "Brand","brand",
]

# ─────────────────────────────────────────────
# MASTER FIELD EXTRACTOR
# ─────────────────────────────────────────────
def extract_field_from_body(body: str, label: str) -> str:
    """
    Handles ALL Outlook table formats:
      Format A  →  label : value
      Format B  →  label - value
      Format C  →  label \\t value
      Format D  →  label     value   (spaces only)
      Format E  →  label\\nvalue     (next line)
      Format F  →  CampaignName value (no space in label)
    """
    if not body:
        return ""

    body_clean  = str(body).replace("\r\n", "\n").replace("\r", "\n")
    lines       = [ln.strip() for ln in body_clean.split("\n")]
    label_norm  = re.sub(r'\s+', ' ', label.strip()).lower()
    label_nospace = label_norm.replace(' ', '')

    for i, line in enumerate(lines):
        if not line:
            continue
        line_norm    = re.sub(r'\s+', ' ', line.strip())
        line_lower   = line_norm.lower()
        line_nospace = line_lower.replace(' ', '')

        # Strategy 1: startswith + optional separator
        for lbl in (label_norm, label_nospace):
            if line_lower.startswith(lbl):
                remainder = line_norm[len(lbl):].strip()
                remainder = remainder.lstrip(":- \t")
                if remainder:
                    return remainder.strip()

        # Strategy 2: separator-based split
        for sep in [':', '-', '\t']:
            if sep in line_norm:
                parts = line_norm.split(sep, 1)
                left  = re.sub(r'\s+', ' ', parts[0].strip()).lower()
                right = parts[1].strip()
                if (left == label_norm or
                        left.replace(' ', '') == label_nospace):
                    if right:
                        return right

        # Strategy 3: label on its own line, value on next line
        if (line_lower == label_norm or
                line_nospace == label_nospace):
            for j in range(i + 1, min(i + 4, len(lines))):
                next_ln = lines[j].strip()
                if next_ln:
                    next_lower = next_ln.lower()
                    looks_like_label = any(
                        next_lower.startswith(kw) for kw in [
                            'campaign','project','client','oli','cid',
                            'company','account','asset','country','email',
                        ])
                    if not looks_like_label:
                        return next_ln
                    break

    return ""


def search_campaign_in_body(body: str) -> str:
    """Priority: Campaign Name → Campaign → Project Name → Project → regex."""
    if not body:
        return ""
    for label in _CAMPAIGN_LABEL_PRIORITY:
        val = extract_field_from_body(body, label)
        if val and is_valid_campaign(val):
            return val.strip()
    # Broad regex sweep
    for pat in [
        r"Campaign\s*Name\s*[=:\-\t]+\s*(.+?)(?:\n|$)",
        r"Project\s*Name\s*[=:\-\t]+\s*(.+?)(?:\n|$)",
        r"(?<![A-Za-z])Campaign\s*[=:\-\t]+\s*(.+?)(?:\n|$)",
        r"(?<![A-Za-z])Project\s*[=:\-\t]+\s*(.+?)(?:\n|$)",
    ]:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            candidate = m.group(1).strip().split("\n")[0].strip()
            if is_valid_campaign(candidate):
                return candidate
    return ""


def search_oli_in_body(body: str) -> str:
    if not body:
        return ""
    for label in _OLI_LABEL_PRIORITY:
        val = extract_field_from_body(body, label)
        if val:
            cleaned = normalize_oli(val)
            if cleaned:
                return cleaned
    m = re.search(
        r"(?:OLI\s*/\s*CID|OLI|CID)\s*[=:\-\t]+\s*([\w\-]+)",
        body, re.IGNORECASE)
    if m:
        return normalize_oli(m.group(1))
    return ""


def search_client_in_body(body: str) -> str:
    if not body:
        return ""
    for label in _CLIENT_LABEL_PRIORITY:
        val = extract_field_from_body(body, label)
        if val and len(val.strip()) > 1:
            return val.strip()
    return ""


def parse_email_body(body: str) -> dict:
    """
    Returns:
      campaign : real name or "" (never "Unknown")
      oli      : normalised  or "" (never "0000")
      client   : name        or ""
      source   : which label matched (for logging)
    """
    if not body:
        return {"campaign": "", "oli": "", "client": "", "source": ""}

    campaign = search_campaign_in_body(body)
    oli      = search_oli_in_body(body)
    client   = search_client_in_body(body)

    source = ""
    if campaign:
        for lbl in ["Campaign Name", "Campaign"]:
            if extract_field_from_body(body, lbl) == campaign:
                source = lbl; break
        if not source:
            for lbl in ["Project Name", "Project"]:
                if extract_field_from_body(body, lbl) == campaign:
                    source = "Project Name (→ used as Campaign)"; break
        if not source:
            source = "regex"

    return {"campaign": campaign, "oli": oli,
            "client": client, "source": source}


def extract_client_name(body: str, subject: str = "") -> str:
    val = search_client_in_body(body)
    if val and len(val) > 1:
        return val
    m = re.search(r"(?:from|for|re:?\s)\s+([A-Z][\w\s&]+)",
                  subject, re.IGNORECASE)
    if m:
        v = m.group(1).strip()
        if v and 1 < len(v) < 60:
            return v
    return "Unknown"

# ─────────────────────────────────────────────
# GARTNER HELPERS
# ─────────────────────────────────────────────
def normalize_client_name(name):
    txt = str(name).strip().lower()
    txt = re.sub(r'[^\w\s]', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt

def is_gartner_client(client_name: str) -> bool:
    normalized = normalize_client_name(client_name)
    return normalized in {"gartner", "gartner inc"}

def get_segment_short_name(segment):
    seg_cleaned = _PORTAL_ILLEGAL_RE.sub('', str(segment)).strip()
    seg_cleaned = _safe_re.sub('', seg_cleaned).strip()
    seg_cleaned = _nonprint_re.sub('', seg_cleaned)
    parts = [p for p in seg_cleaned.split('_') if p]
    if len(parts) <= 1:
        parts = [p for p in _spaces_re.split(seg_cleaned) if p]
    if not parts:
        return "UnknownSegment"
    if len(parts) <= 2:
        selected = parts
    else:
        first_two = parts[:2]
        last_part = parts[-1]
        selected  = first_two if last_part in first_two else first_two + [last_part]
    seg_part = '_'.join(selected)
    seg_part = _MULTI_UNDERSCORE.sub('_', seg_part).strip('_')
    seg_part = _nonprint_re.sub('', seg_part)
    return seg_part[:60] if seg_part else "UnknownSegment"

def get_gartner_campaign_name_for_file(campaign_from_mail):
    raw  = _PORTAL_ILLEGAL_RE.sub('', str(campaign_from_mail)).strip()
    raw  = _safe_re.sub('', raw).strip()
    raw  = _nonprint_re.sub('', raw)
    words = [w for w in _spaces_re.split(raw) if w]
    if len(words) == 1 and '_' in words[0]:
        parts    = [p for p in words[0].split('_') if p]
        selected = parts[:2]
    else:
        selected = words[:2]
    if not selected:
        return "UnknownCampaign"
    result = '_'.join(selected)
    result = _MULTI_UNDERSCORE.sub('_', result).strip('_')
    result = _nonprint_re.sub('', result)
    return result[:80] if result else "UnknownCampaign"

# ─────────────────────────────────────────────
# FILENAME BUILDERS  ← ALL SANITIZED
# ─────────────────────────────────────────────
def format_filename_gartner(campaign_from_mail, segment_from_file,
                             asset, zone, count):
    d          = datetime.now().strftime("%d_%m_%y")
    camp_part  = sanitize_for_portal(
        get_gartner_campaign_name_for_file(campaign_from_mail), 40)
    seg_part   = sanitize_for_portal(
        get_segment_short_name(segment_from_file), 30)
    asset_part = sanitize_for_portal(
        get_asset_name_no_stopwords(asset, max_words=5), 40)
    name = f"WP_{camp_part}_{seg_part}_{asset_part}_{zone}_{count}_{d}"
    return sanitize_for_portal(name) + ".csv"


def format_filename(campaign, oli, asset, tag, count,
                    is_b2bmg=False, b2bmg_oli=None):
    """
    Build CSV filename — fully sanitized, portal-safe.
    OLI omitted when empty. No '0000' ever.
    """
    d = datetime.now().strftime("%d_%m_%y")

    # Sanitize asset
    asset_part = sanitize_for_portal(
        get_asset_name_no_stopwords(str(asset), max_words=5), 50)
    if not asset_part:
        asset_part = "Unknown_Asset"

    if is_b2bmg:
        oli_part = normalize_oli(b2bmg_oli) or normalize_oli(oli) or ""
        if oli_part:
            name = f"WP_B2BMG_{oli_part}_{asset_part}_{tag}_{count}_{d}"
        else:
            name = f"WP_B2BMG_{asset_part}_{tag}_{count}_{d}"
        return sanitize_for_portal(name) + ".csv"

    # Sanitize campaign first word
    camp_raw  = get_first_word(campaign) if is_valid_campaign(campaign) else "WP"
    camp_part = sanitize_for_portal(camp_raw, 40)
    oli_clean = normalize_oli(oli)

    if oli_clean:
        name = f"WP_{camp_part}_{oli_clean}_{asset_part}_{tag}_{count}_{d}"
    else:
        name = f"WP_{camp_part}_{asset_part}_{tag}_{count}_{d}"

    return sanitize_for_portal(name) + ".csv"

# ─────────────────────────────────────────────
# B2BMG HELPERS
# ─────────────────────────────────────────────
def extract_b2bmg_password(subject):
    if not re.search(r'B2BMG', subject, re.IGNORECASE):
        return None
    m = re.search(r'\b(\d{4})\b', subject)
    return f"B2BMG{m.group(1)}" if m else None

def extract_b2bmg_oli(subject):
    if not re.search(r'B2BMG', subject, re.IGNORECASE):
        return None
    m = re.search(r'\b(\d{4})\b', subject)
    return m.group(1) if m else None

# ─────────────────────────────────────────────
# OPEN PASSWORD-PROTECTED EXCEL
# ─────────────────────────────────────────────
def open_protected_excel(file_path, password):
    try:
        import msoffcrypto, io
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            if not office_file.is_encrypted():
                return None, False
            decrypted = io.BytesIO()
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return decrypted, True
    except Exception:
        return None, True

# ─────────────────────────────────────────────
# CAMPAIGN FOLDERS
# ─────────────────────────────────────────────
def create_campaign_folders(base_save_folder, campaign_name):
    safe_camp = safe_folder_name(campaign_name, max_len=60)
    if not safe_camp:
        safe_camp = "Unknown_Campaign"
    campaign_folder    = os.path.join(base_save_folder, safe_camp)
    attachments_folder = os.path.join(campaign_folder, "Attachments")
    csv_folder         = os.path.join(campaign_folder, "CSV")
    try:
        os.makedirs(attachments_folder, exist_ok=True)
        os.makedirs(csv_folder,         exist_ok=True)
    except OSError:
        ts                 = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_camp          = f"Campaign_{ts}"
        campaign_folder    = os.path.join(base_save_folder, safe_camp)
        attachments_folder = os.path.join(campaign_folder, "Attachments")
        csv_folder         = os.path.join(campaign_folder, "CSV")
        os.makedirs(attachments_folder, exist_ok=True)
        os.makedirs(csv_folder,         exist_ok=True)
    return {
        "campaign_folder"   : campaign_folder,
        "attachments_folder": attachments_folder,
        "csv_folder"        : csv_folder,
    }

# ─────────────────────────────────────────────
# GET COLUMNS FROM FILE
# ─────────────────────────────────────────────
def get_columns_from_file(file_path, file_password=None):
    try:
        if file_path.lower().endswith('.csv'):
            return pd.read_csv(file_path, nrows=0).columns.tolist()
        if file_password:
            buf, was_enc = open_protected_excel(file_path, file_password)
            if was_enc and buf:
                xl  = pd.ExcelFile(buf)
                dfs = [xl.parse(s, nrows=0) for s in xl.sheet_names]
                return pd.concat(dfs, ignore_index=True).columns.tolist()
        xl  = pd.ExcelFile(file_path)
        dfs = [xl.parse(s, nrows=0) for s in xl.sheet_names]
        return pd.concat(dfs, ignore_index=True).columns.tolist()
    except Exception:
        return []

# ─────────────────────────────────────────────
# GET HIGHLIGHTED DATA
# ─────────────────────────────────────────────
def get_highlighted_data(file_path, email_col,
                          file_password=None, log_func=None):
    def _log(m):
        if log_func: log_func(m)

    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        return pd.read_csv(file_path)

    wb = None
    if file_password:
        buf, was_encrypted = open_protected_excel(file_path, file_password)
        if was_encrypted and buf:
            try:
                wb = openpyxl.load_workbook(buf, data_only=True, read_only=False)
                _log("   🔐 Workbook decrypted successfully")
            except Exception as ex:
                _log(f"   ⚠️  Decrypt OK but openpyxl failed: {ex}")
                wb = None
        elif was_encrypted and not buf:
            _log("   ❌ Wrong password — trying without password")

    if wb is None:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as ex:
            _log(f"   ❌ Cannot open workbook: {ex}")
            return pd.DataFrame()

    YELLOWS    = {"FFFF00","FFFFFF00","FFFFE0","FFF2CC"}
    all_frames = []
    _log(f"   📑 Sheets found: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        try:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
            headers   = [c.value for c in first_row]
        except StopIteration:
            _log(f"   ⚠️  Sheet '{sheet_name}' empty — skipping")
            continue

        ei = None
        for idx, h in enumerate(headers):
            if h is not None and str(h).strip() == email_col:
                ei = idx; break
        if ei is None:
            for idx, h in enumerate(headers):
                if h is not None and str(h).strip().lower() == email_col.lower():
                    ei = idx; break
        if ei is None:
            _log(f"   ⚠️  Sheet '{sheet_name}': '{email_col}' not found — skipping")
            continue

        sheet_data = []
        for row in sheet.iter_rows(min_row=2):
            cell      = row[ei]
            fill      = cell.fill
            rgb       = str(fill.start_color.rgb).upper()
            idx_color = fill.start_color.index
            is_yellow = (
                any(y in rgb for y in YELLOWS) or
                idx_color == 6 or
                (fill.start_color.type == 'theme' and
                 fill.start_color.theme == 4)
            )
            if is_yellow:
                sheet_data.append([c.value for c in row])

        if sheet_data:
            df_sheet = pd.DataFrame(sheet_data, columns=headers)
            all_frames.append(df_sheet)
            _log(f"   ✅ Sheet '{sheet_name}': {len(sheet_data)} highlighted rows")
        else:
            _log(f"   ○  Sheet '{sheet_name}': 0 highlighted rows")

    wb.close()

    if not all_frames:
        _log("   ⚠️  No highlighted rows found in any sheet")
        return pd.DataFrame()

    combined = pd.concat(all_frames, ignore_index=True)
    _log(f"   📊 Total combined rows: {len(combined)}")
    return combined

def auto_map_columns(cols):
    lc  = {str(c).strip().lower(): str(c).strip() for c in cols}
    res = {}
    for field, cands in AUTO_MAP.items():
        res[field] = next((lc[c] for c in cands if c in lc), "")
    return res

def fix_column(c):
    return str(c).replace("'", "").replace(",", "").strip()

# ─────────────────────────────────────────────
# PORTAL LOGIN
# ─────────────────────────────────────────────
def portal_login(page, username, password, log_func=None):
    def _log(m):
        if log_func: log_func(m)

    _log("🌐 Opening portal login...")
    page.goto(PORTAL_CONFIG["login_url"])
    page.wait_for_load_state("domcontentloaded")
    time.sleep(1)

    filled = False
    for sel in PORTAL_CONFIG["email_sel"]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1500):
                el.click(); el.fill(""); el.fill(username)
                filled = True; break
        except: continue
    if not filled:
        try:   page.locator("input:visible").nth(0).fill(username); filled = True
        except: pass
    if not filled: raise Exception("Cannot find email field.")
    _log("✅ Email filled")

    filled = False
    for sel in PORTAL_CONFIG["pass_sel"]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1500):
                el.click(); el.fill(""); el.fill(password)
                filled = True; break
        except: continue
    if not filled:
        try:   page.locator("input:visible").nth(1).fill(password); filled = True
        except: pass
    if not filled: raise Exception("Cannot find password field.")
    _log("✅ Password filled")

    clicked = False
    for sel in PORTAL_CONFIG["signin_sel"]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=1500):
                btn.click(); clicked = True; break
        except: continue
    if not clicked: page.keyboard.press("Enter")
    _log("✅ Sign In clicked")

    page.wait_for_load_state("networkidle", timeout=25000)
    time.sleep(WAIT_AFTER_LOGIN)
    if "login" in page.url.lower() or "signin" in page.url.lower():
        raise Exception("Login failed — still on login page.")
    _log("✅ Login successful!")
    return True


def click_create_new_list(page, log_func=None):
    def _log(m):
        if log_func: log_func(m)

    _log("🖱️  'Create New List'...")
    try:   page.wait_for_load_state("networkidle", timeout=10000)
    except: pass
    time.sleep(WAIT_AFTER_NAV)

    strategies = [
        lambda: page.get_by_role("button", name=re.compile(r"Create New List", re.IGNORECASE)).first,
        lambda: page.get_by_role("link",   name=re.compile(r"Create New List", re.IGNORECASE)).first,
        lambda: page.get_by_text("Create New List", exact=True).first,
        lambda: page.get_by_text(re.compile(r"Create New List", re.IGNORECASE)).first,
        lambda: page.get_by_text(re.compile(r"New List", re.IGNORECASE)).first,
        lambda: page.locator('button:has-text("Create New List")').first,
        lambda: page.locator('a:has-text("Create New List")').first,
        lambda: page.locator('button:has-text("New List")').first,
        lambda: page.locator('[data-action*="create" i]').first,
        lambda: page.locator('[id*="create" i]').first,
    ]
    for fn in strategies:
        try:
            el = fn()
            if el.is_visible(timeout=2000):
                el.scroll_into_view_if_needed(); el.click(timeout=4000)
                _log("✅ 'Create New List' clicked"); return True
        except: continue

    try:
        for el in page.locator("button,a,[role='button'],[role='link']").all():
            try:
                txt = (el.inner_text() or "").strip().lower()
                if ("create" in txt and "list" in txt) or "new list" in txt:
                    el.scroll_into_view_if_needed(); el.click(timeout=4000)
                    _log("✅ Clicked via scan"); return True
            except: continue
    except: pass
    raise Exception("❌ Could not find 'Create New List' button.")


def click_single_list_from_dropdown(page, log_func=None):
    def _log(m):
        if log_func: log_func(m)

    _log("🖱️  'Single List (Classic)'...")
    time.sleep(WAIT_AFTER_CREATE_CLICK)

    strategies = [
        lambda: page.get_by_role("menuitem", name=re.compile(r"Single List", re.IGNORECASE)).first,
        lambda: page.get_by_role("option",   name=re.compile(r"Single List", re.IGNORECASE)).first,
        lambda: page.get_by_text("Single List (Classic)", exact=True).first,
        lambda: page.get_by_text(re.compile(r"Single List (Classic)", re.IGNORECASE)).first,
        lambda: page.locator('li:has-text("Single List (Classic)")').first,
        lambda: page.locator('a:has-text("Single List (Classic)")').first,
        lambda: page.locator('.dropdown-item:has-text("Single List (Classic)")').first,
        lambda: page.locator('[role="menuitem"]:has-text("Single List (Classic)")').first,
        lambda: page.locator('ul li:has-text("Single List (Classic)")').first,
    ]
    for fn in strategies:
        try:
            el = fn()
            if el.is_visible(timeout=2500):
                el.scroll_into_view_if_needed(); el.click(timeout=4000)
                _log("✅ 'Single List (Classic)' clicked"); return True
        except: continue

    try:
        for el in page.locator("li,a,[role='menuitem'],[role='option']").all():
            try:
                txt = (el.inner_text() or "").strip().lower()
                if "single" in txt:
                    el.scroll_into_view_if_needed(); el.click(timeout=4000)
                    _log("✅ Clicked via scan"); return True
            except: continue
    except: pass
    raise Exception("❌ Could not find 'Single List (Classic)' option.")


def fill_list_name(page, list_name, log_func=None):
    def _log(m):
        if log_func: log_func(m)

    _log(f"✏️  Filling List Name: '{list_name}'")
    try:   page.wait_for_load_state("networkidle", timeout=10000)
    except: pass
    time.sleep(WAIT_AFTER_SINGLE_LIST)

    for sel in [
        'input[placeholder="Type Title of your subscriber list"]',
        'input[placeholder="Type Title of your subscriber list "]',
        "input[placeholder='Type Title of your subscriber list']",
    ]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=3000):
                el.scroll_into_view_if_needed(); el.click()
                el.select_text(); el.fill(""); el.fill(list_name)
                if el.input_value() == list_name:
                    _log("✅ List Name filled"); return True
                el.triple_click(); el.type(list_name, delay=50)
                _log("✅ List Name typed"); return True
        except: continue

    for sel in [
        'input[placeholder*="Type Title" i]',
        'input[placeholder*="subscriber list" i]',
        'input[placeholder*="Title" i]',
        'input[placeholder*="title" i]',
        'input[placeholder*="list name" i]',
        'input[placeholder*="name" i]',
    ]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2500):
                el.scroll_into_view_if_needed(); el.click()
                el.triple_click(); el.fill(""); el.fill(list_name)
                _log("✅ List Name filled"); return True
        except: continue

    for sel in [
        'input[name*="title" i]','input[name*="name" i]','input[name*="list" i]',
        'input[id*="title" i]',  'input[id*="name" i]', 'input[id*="list" i]',
    ]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2500):
                el.scroll_into_view_if_needed(); el.click()
                el.triple_click(); el.fill(""); el.fill(list_name)
                _log("✅ List Name filled"); return True
        except: continue

    for label_text in ["List Name","list name","Name","Title"]:
        try:
            lbl = page.get_by_label(label_text)
            if lbl.is_visible(timeout=1500):
                lbl.click(); lbl.triple_click(); lbl.fill(""); lbl.fill(list_name)
                _log("✅ List Name filled via label"); return True
        except: continue

    try:
        for inp in page.locator('input[type="text"]').all():
            try:
                if inp.is_visible(timeout=1500):
                    inp.scroll_into_view_if_needed(); inp.click()
                    inp.triple_click(); inp.fill(""); inp.fill(list_name)
                    _log("✅ List Name filled (fallback)"); return True
            except: continue
    except: pass

    raise Exception(f"❌ Could not fill 'List Name' with '{list_name}'.")

# ─────────────────────────────────────────────
# UI COMPONENTS
# ─────────────────────────────────────────────
class ScrollableFrame(tk.Frame):
    def __init__(self, parent, bg=Theme.BG_BASE, **kw):
        super().__init__(parent, bg=bg, **kw)
        self._canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self._sb     = tk.Scrollbar(self, orient="vertical",
                                    command=self._canvas.yview,
                                    width=6, relief='flat')
        self.inner   = tk.Frame(self._canvas, bg=bg)
        self._win_id = self._canvas.create_window(
            (0, 0), window=self.inner, anchor="nw")
        self._canvas.configure(yscrollcommand=self._sb.set)
        self._sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self.inner.bind("<Configure>",
                        lambda e: self._canvas.configure(
                            scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>",
                          lambda e: self._canvas.itemconfig(
                              self._win_id, width=e.width))
        self._canvas.bind("<Enter>",
                          lambda e: self._canvas.bind_all(
                              "<MouseWheel>",
                              lambda ev: self._canvas.yview_scroll(
                                  int(-1*(ev.delta/120)), "units")))
        self._canvas.bind("<Leave>",
                          lambda e: self._canvas.unbind_all("<MouseWheel>"))


class DecorativeCanvas(tk.Canvas):
    def __init__(self, parent, **kw):
        super().__init__(parent, highlightthickness=0, **kw)
        self.bind("<Configure>", self._draw)

    def _draw(self, e=None):
        self.delete("deco")
        w, h = self.winfo_width(), self.winfo_height()
        for cx, cy, r, col in [
            (w-80,40,120,"#EDE9FE"),(w-200,-30,90,"#DBEAFE"),
            (60,h-60,100,"#FEF3C7"),(w-40,h-100,70,"#D1FAE5"),
            (-40,80,80,"#FCE7F3"),(w//2,-20,60,"#E0E7FF"),
        ]:
            self.create_oval(cx-r,cy-r,cx+r,cy+r,
                             fill=col,outline="",tags="deco")
        for x in range(40,w,50):
            for y in range(40,h,50):
                self.create_oval(x,y,x+3,y+3,
                                 fill="#E5E1DA",outline="",tags="deco")
        self.tag_lower("deco")


class PortalCard(tk.Frame):
    def __init__(self, parent, title="", icon="", **kw):
        super().__init__(parent, bg=Theme.BG_CARD, bd=0,
                         highlightbackground=Theme.BORDER_CARD,
                         highlightthickness=1, **kw)
        if title:
            h = tk.Frame(self, bg=Theme.BG_CARD, pady=7, padx=12)
            h.pack(fill='x')
            if icon:
                tk.Label(h, text=icon, fg=Theme.FG_ACCENT,
                         bg=Theme.BG_CARD,
                         font=Theme.FONT_ICON).pack(side='left', padx=(0,6))
            tk.Label(h, text=title, fg=Theme.FG_PRIMARY,
                     bg=Theme.BG_CARD,
                     font=Theme.FONT_SUBTITLE).pack(side='left')
            tk.Frame(self, bg=Theme.BORDER, height=1).pack(fill='x', padx=12)
        self.body = tk.Frame(self, bg=Theme.BG_CARD, padx=12, pady=8)
        self.body.pack(fill='both', expand=True)


class StatsCard(tk.Frame):
    def __init__(self, parent, icon="", title="", value="0",
                 accent=Theme.BG_ACCENT, accent_bg=Theme.BG_ACCENT_LIGHT, **kw):
        super().__init__(parent, bg=Theme.BG_CARD, padx=12, pady=8,
                         highlightbackground=Theme.BORDER_CARD,
                         highlightthickness=1, **kw)
        row = tk.Frame(self, bg=Theme.BG_CARD); row.pack(fill='x')
        ic  = tk.Frame(row, bg=accent_bg, width=32, height=32)
        ic.pack_propagate(False); ic.pack(side='left')
        tk.Label(ic, text=icon, bg=accent_bg, fg=accent,
                 font=(Theme.FONT_FAMILY,12)).place(
                     relx=0.5, rely=0.5, anchor='center')
        rb = tk.Frame(row, bg=Theme.BG_CARD)
        rb.pack(side='left', padx=(8,0))
        self.val = tk.Label(rb, text=value, fg=Theme.FG_PRIMARY,
                            bg=Theme.BG_CARD,
                            font=(Theme.FONT_FAMILY,16,"bold"))
        self.val.pack(anchor='w')
        tk.Label(rb, text=title, fg=Theme.FG_SECONDARY,
                 bg=Theme.BG_CARD,
                 font=Theme.FONT_TINY).pack(anchor='w')

    def set_value(self, v): self.val.config(text=str(v))


class PortalButton(tk.Frame):
    def __init__(self, parent, text="", icon="",
                 bg_color=Theme.BG_BLUE_BTN, fg_color=Theme.FG_WHITE,
                 hover_color=None, command=None, font=None, **kw):
        try:    pbg = parent.cget("bg")
        except: pbg = Theme.BG_BASE
        super().__init__(parent, bg=pbg, **kw)
        self._bg  = bg_color; self._fg  = fg_color
        self._hov = hover_color or self._lghtn(bg_color, 20)
        self._cmd = command; self._fnt = font or Theme.FONT_BODY
        disp      = f"{icon} {text}" if icon else text
        self._out = tk.Frame(self, bg=bg_color, padx=2, pady=2)
        self._out.pack(fill='x')
        self._lbl = tk.Label(self._out, text=disp, bg=bg_color, fg=fg_color,
                             font=self._fnt, cursor="hand2", padx=14, pady=7)
        self._lbl.pack(fill='x')
        self._lbl.bind("<Enter>",           self._ent)
        self._lbl.bind("<Leave>",           self._lev)
        self._lbl.bind("<Button-1>",        self._clk)
        self._lbl.bind("<ButtonRelease-1>", self._rel)

    def _ent(self, e): self._out.config(bg=self._hov); self._lbl.config(bg=self._hov)
    def _lev(self, e): self._out.config(bg=self._bg);  self._lbl.config(bg=self._bg)
    def _clk(self, e):
        d = self._drkn(self._bg, 20)
        self._out.config(bg=d); self._lbl.config(bg=d)
    def _rel(self, e):
        self._out.config(bg=self._bg); self._lbl.config(bg=self._bg)
        if self._cmd: self._cmd()

    @staticmethod
    def _lghtn(hc, a):
        h = hc.lstrip('#')
        return "#{:02x}{:02x}{:02x}".format(
            min(255,int(h[0:2],16)+a),
            min(255,int(h[2:4],16)+a),
            min(255,int(h[4:6],16)+a))

    @staticmethod
    def _drkn(hc, a):
        h = hc.lstrip('#')
        return "#{:02x}{:02x}{:02x}".format(
            max(0,int(h[0:2],16)-a),
            max(0,int(h[2:4],16)-a),
            max(0,int(h[4:6],16)-a))


class MappingRow(tk.Frame):
    def __init__(self, parent, field_name, icon="", required=True, **kw):
        super().__init__(parent, bg=Theme.BG_CARD, **kw)
        self.field_name = field_name
        self.var        = tk.StringVar()
        top             = tk.Frame(self, bg=Theme.BG_CARD)
        top.pack(fill='x', pady=(3,1))
        lbl = f"{icon} {field_name}" if icon else field_name
        tk.Label(top, text=lbl,
                 fg=Theme.FG_PRIMARY if required else Theme.FG_SECONDARY,
                 bg=Theme.BG_CARD, font=Theme.FONT_BODY,
                 anchor='w').pack(side='left')
        if required:
            tk.Label(top, text=" *", fg=Theme.FG_DANGER,
                     bg=Theme.BG_CARD, font=Theme.FONT_SMALL).pack(side='left')
        sf = tk.Frame(top, bg=Theme.BG_CARD); sf.pack(side='right')
        self.sdot = tk.Label(sf, text="●", fg=Theme.FG_MUTED,
                             bg=Theme.BG_CARD, font=(Theme.FONT_FAMILY,6))
        self.sdot.pack(side='left', padx=(0,2))
        self.stxt = tk.Label(sf, text="", fg=Theme.FG_MUTED,
                             bg=Theme.BG_CARD, font=Theme.FONT_TINY)
        self.stxt.pack(side='left')
        dd = tk.Frame(self, bg=Theme.BG_INPUT,
                      highlightbackground=Theme.BORDER,
                      highlightcolor=Theme.BORDER_FOCUS,
                      highlightthickness=1)
        dd.pack(fill='x', pady=(0,3))
        self.om = tk.OptionMenu(dd, self.var, "")
        self.om.config(bg=Theme.BG_INPUT, fg=Theme.FG_INPUT,
                       font=Theme.FONT_BODY, relief='flat',
                       highlightthickness=0, bd=3,
                       activebackground=Theme.BG_HOVER,
                       activeforeground=Theme.FG_PRIMARY,
                       anchor='w', cursor="hand2")
        self.om["menu"].config(bg=Theme.BG_WHITE, fg=Theme.FG_PRIMARY,
                               activebackground=Theme.BG_ACCENT_LIGHT,
                               activeforeground=Theme.FG_ACCENT,
                               font=Theme.FONT_BODY, relief='flat', bd=3)
        self.om.pack(fill='x')
        self._is_manual = False
        self.var.trace_add("write", self._on_var_change)

    def _on_var_change(self, *args):
        self._is_manual = True
        self.sdot.config(fg="#F97316")
        self.stxt.config(text="manual", fg="#F97316")

    def update_options(self, cols, auto=""):
        self._is_manual = False
        m = self.om["menu"]; m.delete(0,"end")
        for c in cols:
            m.add_command(label=c, command=lambda v=c: self._manual_set(v))
        if auto:
            self.var.set(auto)
            self._is_manual = False
            self.sdot.config(fg=Theme.FG_SUCCESS)
            self.stxt.config(text="auto", fg=Theme.FG_SUCCESS)
        else:
            self.var.set("")
            self._is_manual = False
            self.sdot.config(fg=Theme.FG_MUTED)
            self.stxt.config(text="not mapped", fg=Theme.FG_MUTED)

    def _manual_set(self, v):
        self._is_manual = True
        self.var.set(v)
        self.sdot.config(fg="#F97316")
        self.stxt.config(text="manual ✓", fg="#F97316")

    def get_value(self):  return self.var.get().strip()
    def is_manual(self):  return self._is_manual


class LogConsole(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=Theme.BG_CONSOLE, **kw)
        hdr = tk.Frame(self, bg="#282A36", padx=12, pady=5)
        hdr.pack(fill='x')
        dots = tk.Frame(hdr, bg="#282A36"); dots.pack(side='left')
        for c in ["#FF5F56","#FFBD2E","#27C93F"]:
            tk.Label(dots, text="●", fg=c, bg="#282A36",
                     font=(Theme.FONT_FAMILY,7)).pack(side='left', padx=2)
        tk.Label(hdr, text=" Console Output", fg="#6272A4",
                 bg="#282A36", font=Theme.FONT_SMALL).pack(side='left', padx=(6,0))
        clr = tk.Label(hdr, text="Clear", fg="#6272A4",
                       bg="#282A36", font=Theme.FONT_SMALL, cursor="hand2")
        clr.pack(side='right')
        clr.bind("<Button-1>", lambda e: self.clear())
        clr.bind("<Enter>",    lambda e: clr.config(fg="#F8F8F2"))
        clr.bind("<Leave>",    lambda e: clr.config(fg="#6272A4"))
        tf = tk.Frame(self, bg="#1E1E2E"); tf.pack(fill='both', expand=True)
        self.text = tk.Text(tf, bg="#1E1E2E", fg="#CDD6F4",
                            font=Theme.FONT_MONO, relief='flat', bd=10,
                            insertbackground="#F8F8F2", wrap='word',
                            selectbackground="#44475A", cursor="arrow")
        sb = tk.Scrollbar(tf, command=self.text.yview, bg="#282A36",
                          troughcolor="#1E1E2E", width=7, relief='flat')
        self.text.config(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        self.text.pack(fill='both', expand=True)
        self.text.tag_config("success", foreground="#50FA7B")
        self.text.tag_config("error",   foreground="#FF5555")
        self.text.tag_config("warning", foreground="#F1FA8C")
        self.text.tag_config("info",    foreground="#8BE9FD")
        self.text.tag_config("header",  foreground="#BD93F9",
                             font=(Theme.FONT_FAMILY,10,"bold"))
        self.text.tag_config("muted",   foreground="#6272A4")
        self.text.tag_config("time",    foreground="#44475A")
        self.text.tag_config("manual",  foreground="#F97316")
        self.text.tag_config("project", foreground="#FFB86C")

    def log(self, msg, tag=None):
        if tag is None:
            if   "✅" in msg: tag = "success"
            elif "❌" in msg: tag = "error"
            elif "⚠"  in msg: tag = "warning"
            elif any(s in msg for s in ["📊","🚀"]): tag = "header"
            elif any(s in msg for s in ["💾","📂"]): tag = "info"
            else: tag = "muted"
        ts = datetime.now().strftime("%H:%M:%S")
        self.text.insert(tk.END, f" {ts}  ", "time")
        self.text.insert(tk.END, msg, tag)
        if not msg.endswith("\n"):
            self.text.insert(tk.END, "\n")
        self.text.see(tk.END)

    def clear(self): self.text.delete("1.0", tk.END)


class SectionLabel(tk.Frame):
    def __init__(self, parent, text, icon="", color=Theme.BG_ACCENT, **kw):
        super().__init__(parent, bg=Theme.BG_BASE, **kw)
        bar = tk.Frame(self, bg=color, width=3)
        bar.pack(side='left', fill='y')
        tk.Label(self,
                 text=f" {icon} {text}" if icon else f" {text}",
                 fg=color, bg=Theme.BG_BASE,
                 font=(Theme.FONT_FAMILY,8,"bold")).pack(side='left', pady=4)

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class SmartLeadApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SmartLead Portal")
        self.root.geometry("1400x920")
        self.root.configure(bg=Theme.BG_BASE)
        self.root.minsize(1200, 800)

        self.sys_user        = get_system_user_info()
        self.user_name_var   = tk.StringVar(value=self.sys_user["employee_id"])
        self.file_path_var   = tk.StringVar()
        self.save_folder_var = tk.StringVar()
        self.username_var    = tk.StringVar()
        self.password_var    = tk.StringVar()
        self.is_logged_in    = False

        self.outlook_mail_jobs = []
        self.mapping_rows      = {}
        self.stats_cards       = {}
        self.tracking_records  = []

        self._build_ui()
        self._welcome()

    def _launch_browser(self, pw):
        chrome_path = find_chrome_path()
        if chrome_path:
            self.log(f"🌐 Using: {os.path.basename(chrome_path)}", "info")
            try:
                return pw.chromium.launch(
                    executable_path=chrome_path, headless=False,
                    args=["--no-sandbox","--disable-dev-shm-usage",
                          "--disable-gpu","--disable-extensions",
                          "--start-maximized"])
            except Exception as e:
                self.log(f"⚠️  System browser failed: {e}", "warning")
        try:
            return pw.chromium.launch(headless=False)
        except Exception as e:
            raise Exception(
                f"❌ No browser found!\nhttps://www.google.com/chrome/\n{e}")

    def _build_ui(self):
        bg   = DecorativeCanvas(self.root, bg=Theme.BG_BASE)
        bg.place(x=0, y=0, relwidth=1, relheight=1)
        wrap = tk.Frame(self.root, bg=Theme.BG_BASE)
        wrap.place(x=0, y=0, relwidth=1, relheight=1)
        self._build_header(wrap)
        self._build_stats_row(wrap)
        self._build_main(wrap)

    def _build_header(self, parent):
        hdr = tk.Frame(parent, bg=Theme.BG_WHITE, pady=9, padx=18,
                       highlightbackground=Theme.BORDER_LIGHT,
                       highlightthickness=1)
        hdr.pack(fill='x')
        left = tk.Frame(hdr, bg=Theme.BG_WHITE); left.pack(side='left')
        bar  = tk.Frame(left, bg=Theme.BG_ACCENT, width=4, height=30)
        bar.pack(side='left', padx=(0,10)); bar.pack_propagate(False)
        tb = tk.Frame(left, bg=Theme.BG_WHITE); tb.pack(side='left')
        tk.Label(tb, text="SmartLead Portal", fg=Theme.FG_PRIMARY,
                 bg=Theme.BG_WHITE, font=Theme.FONT_TITLE).pack(anchor='w')
        tk.Label(tb, text="Automated Lead Processing & Upload  |  Issues? Contact WFM – Omkar",
                 fg=Theme.FG_MUTED, bg=Theme.BG_WHITE,
                 font=Theme.FONT_TINY).pack(anchor='w')
        right = tk.Frame(hdr, bg=Theme.BG_WHITE); right.pack(side='right')
        self.login_frame = tk.Frame(right, bg=Theme.BG_CARD,
                                    highlightbackground=Theme.BORDER,
                                    highlightthickness=1, padx=8, pady=3)
        self.login_frame.pack(side='right', padx=(8,0))
        self.login_dot = tk.Label(self.login_frame, text="●",
                                  fg=Theme.FG_MUTED, bg=Theme.BG_CARD,
                                  font=(Theme.FONT_FAMILY,7))
        self.login_dot.pack(side='left', padx=(0,4))
        self.login_lbl = tk.Label(self.login_frame, text="Not logged in",
                                  fg=Theme.FG_MUTED, bg=Theme.BG_CARD,
                                  font=Theme.FONT_TINY)
        self.login_lbl.pack(side='left')
        self.status_frame = tk.Frame(right, bg=Theme.BG_SUCCESS_LIGHT,
                                     padx=10, pady=3)
        self.status_frame.pack(side='right', padx=(8,0))
        self.status_dot = tk.Label(self.status_frame, text="●",
                                   fg=Theme.BG_SUCCESS, bg=Theme.BG_SUCCESS_LIGHT,
                                   font=(Theme.FONT_FAMILY,7))
        self.status_dot.pack(side='left', padx=(0,4))
        self.status_lbl = tk.Label(self.status_frame, text="Ready",
                                   fg=Theme.FG_SUCCESS, bg=Theme.BG_SUCCESS_LIGHT,
                                   font=(Theme.FONT_FAMILY,8,"bold"))
        self.status_lbl.pack(side='left')
        self.file_ind = tk.Label(right, text="No file loaded",
                                 fg=Theme.FG_MUTED, bg=Theme.BG_WHITE,
                                 font=Theme.FONT_TINY)
        self.file_ind.pack(side='right', padx=(0,10))

    def _build_stats_row(self, parent):
        sf = tk.Frame(parent, bg=Theme.BG_BASE)
        sf.pack(fill='x', padx=18, pady=(10,6))
        for i,(icon,title,val,acc,abg,key) in enumerate([
            ("📁","Files Loaded",  "0",Theme.BG_ACCENT,Theme.BG_ACCENT_LIGHT,"loaded"),
            ("📋","Rows Processed","0",Theme.BG_INFO,  Theme.BG_INFO_LIGHT,  "rows"),
            ("✅","Uploaded",      "0",Theme.BG_SUCCESS,Theme.BG_SUCCESS_LIGHT,"success"),
            ("⚠️","Failed",        "0",Theme.BG_DANGER,Theme.BG_DANGER_LIGHT, "failed"),
        ]):
            c = StatsCard(sf, icon=icon, title=title, value=val,
                          accent=acc, accent_bg=abg)
            c.pack(side='left', fill='x', expand=True,
                   padx=(0 if i==0 else 4, 0 if i==3 else 4))
            self.stats_cards[key] = c

    def _build_main(self, parent):
        body = tk.Frame(parent, bg=Theme.BG_BASE)
        body.pack(fill='both', expand=True, padx=18, pady=(4,16))
        col1_outer = tk.Frame(body, bg=Theme.BG_BASE, width=300)
        col1_outer.pack(side='left', fill='y', padx=(0,8))
        col1_outer.pack_propagate(False)
        self.left_scroll = ScrollableFrame(col1_outer, bg=Theme.BG_BASE)
        self.left_scroll.pack(fill='both', expand=True)
        L = self.left_scroll.inner
        self._build_user_section(L)
        self._build_credentials_section(L)
        self._build_file_section(L)
        self._build_outlook_section(L)
        self._build_actions_section(L)
        col2 = tk.Frame(body, bg=Theme.BG_BASE, width=420)
        col2.pack(side='left', fill='y', padx=(0,8))
        col2.pack_propagate(False)
        self._build_mapping_card(col2)
        col3 = tk.Frame(body, bg=Theme.BG_BASE)
        col3.pack(side='left', fill='both', expand=True)
        self._build_console(col3)

    def _build_user_section(self, parent):
        SectionLabel(parent, "Logged-in User", "👤",
                     color=Theme.BG_ACCENT).pack(fill='x', pady=(0,4))
        card = PortalCard(parent); card.pack(fill='x', pady=(0,8))
        b    = card.body
        row  = tk.Frame(b, bg=Theme.BG_CARD); row.pack(fill='x', pady=(0,6))
        emp_id   = self.sys_user["employee_id"]
        fullname = self.sys_user["full_name"]
        domain   = self.sys_user["domain"]
        pc       = self.sys_user["pc_name"]
        initials = ("".join(p[0].upper() for p in fullname.split()[:2])
                    if fullname else "?")
        av = tk.Frame(row, bg=Theme.BG_ACCENT, width=36, height=36)
        av.pack_propagate(False); av.pack(side='left', padx=(0,8))
        tk.Label(av, text=initials, bg=Theme.BG_ACCENT, fg=Theme.FG_WHITE,
                 font=(Theme.FONT_FAMILY,11,"bold")).place(
                     relx=0.5, rely=0.5, anchor='center')
        info = tk.Frame(row, bg=Theme.BG_CARD)
        info.pack(side='left', fill='x', expand=True)
        r1 = tk.Frame(info, bg=Theme.BG_CARD); r1.pack(anchor='w')
        tk.Label(r1, text=emp_id, fg=Theme.FG_PRIMARY, bg=Theme.BG_CARD,
                 font=(Theme.FONT_FAMILY,10,"bold")).pack(side='left')
        if domain:
            tk.Label(r1, text=f"@{domain}", fg=Theme.FG_MUTED,
                     bg=Theme.BG_CARD, font=Theme.FONT_TINY).pack(
                         side='left', padx=(3,0))
        subs = []
        if fullname and fullname != emp_id: subs.append(fullname)
        if pc: subs.append(f"PC: {pc}")
        if subs:
            tk.Label(info, text=" | ".join(subs), fg=Theme.FG_MUTED,
                     bg=Theme.BG_CARD, font=Theme.FONT_TINY).pack(anchor='w')
        tk.Label(b, text="🔒  Identity auto-detected from Windows login",
                 fg=Theme.FG_SUCCESS, bg=Theme.BG_SUCCESS_LIGHT,
                 font=Theme.FONT_TINY, padx=6, pady=3).pack(fill='x')

    def _build_credentials_section(self, parent):
        SectionLabel(parent, "Portal Credentials", "🔐",
                     color=Theme.BG_BLUE_BTN).pack(fill='x', pady=(8,4))
        card = PortalCard(parent); card.pack(fill='x', pady=(0,8))
        b    = card.body
        self.login_status_lbl = tk.Label(b, text="● Not connected",
                                         fg=Theme.FG_MUTED, bg=Theme.BG_CARD,
                                         font=Theme.FONT_TINY)
        self.login_status_lbl.pack(anchor='w', pady=(0,6))
        for lbl_txt, var, show in [
            ("Portal Email",    self.username_var, ""),
            ("Portal Password", self.password_var, "●"),
        ]:
            tk.Label(b, text=lbl_txt, fg=Theme.FG_SECONDARY, bg=Theme.BG_CARD,
                     font=Theme.FONT_SMALL, anchor='w').pack(fill='x', pady=(2,0))
            ef = tk.Frame(b, bg=Theme.BG_INPUT,
                          highlightbackground=Theme.BORDER,
                          highlightcolor=Theme.BORDER_FOCUS,
                          highlightthickness=1)
            ef.pack(fill='x', pady=(1,4))
            tk.Entry(ef, textvariable=var, bg=Theme.BG_INPUT,
                     fg=Theme.FG_INPUT, font=Theme.FONT_BODY,
                     relief='flat', bd=5, show=show,
                     insertbackground=Theme.FG_PRIMARY).pack(fill='x')
        note = tk.Frame(b, bg=Theme.BG_WARNING_LIGHT, padx=6, pady=4)
        note.pack(fill='x', pady=(2,6))
        tk.Label(note, text="🔑 These auto-fill the portal\n   Email & Password fields",
                 fg=Theme.FG_WARNING, bg=Theme.BG_WARNING_LIGHT,
                 font=Theme.FONT_TINY, justify='left').pack(anchor='w')
        PortalButton(b, text="Test Login", icon="🔑",
                     bg_color=Theme.BG_TEAL_BTN, font=Theme.FONT_SMALL,
                     command=self.test_login).pack(fill='x')

    def _build_file_section(self, parent):
        SectionLabel(parent, "Data File", "📁",
                     color=Theme.BG_INFO).pack(fill='x', pady=(8,4))
        card = PortalCard(parent); card.pack(fill='x', pady=(0,8))
        b    = card.body
        ef = tk.Frame(b, bg=Theme.BG_INPUT,
                      highlightbackground=Theme.BORDER,
                      highlightcolor=Theme.BORDER_FOCUS, highlightthickness=1)
        ef.pack(fill='x', pady=(0,6))
        tk.Entry(ef, textvariable=self.file_path_var, bg=Theme.BG_INPUT,
                 fg=Theme.FG_INPUT, font=Theme.FONT_TINY, relief='flat',
                 bd=4, insertbackground=Theme.FG_PRIMARY).pack(fill='x')
        browse = tk.Label(b, text="  📁  Browse File  ",
                          fg=Theme.FG_ACCENT, bg=Theme.BG_ACCENT_LIGHT,
                          font=Theme.FONT_SMALL, cursor="hand2", padx=8, pady=5)
        browse.pack(fill='x')
        browse.bind("<Button-1>", lambda e: self.browse_file())
        browse.bind("<Enter>",
                    lambda e: browse.config(bg=Theme.BG_ACCENT, fg=Theme.FG_WHITE))
        browse.bind("<Leave>",
                    lambda e: browse.config(bg=Theme.BG_ACCENT_LIGHT,
                                            fg=Theme.FG_ACCENT))

    def _build_outlook_section(self, parent):
        SectionLabel(parent, "Outlook Sync", "📧",
                     color=Theme.BG_ORANGE_BTN).pack(fill='x', pady=(8,4))
        card = PortalCard(parent); card.pack(fill='x', pady=(0,8))
        b    = card.body
        info = tk.Frame(b, bg=Theme.BG_INFO_LIGHT, padx=6, pady=5)
        info.pack(fill='x', pady=(0,6))
        tk.Label(info,
                 text="📬 Select email(s) in Outlook first,\n"
                      "   then click Sync.\n"
                      "   📣 Campaign Name → Priority 1\n"
                      "   📝 Project Name  → Fallback (Priority 2)\n"
                      "   🔐 B2BMG files auto-unlocked.\n"
                      "   🏢 Gartner: DO + WW by Segment.\n"
                      "   🔢 OLI omitted if not in email.\n"
                      "   ✅ Filenames auto-sanitized for portal.",
                 fg=Theme.FG_SECONDARY, bg=Theme.BG_INFO_LIGHT,
                 font=Theme.FONT_TINY, justify='left').pack(anchor='w')
        PortalButton(b, text="Sync from Outlook", icon="📧",
                     bg_color=Theme.BG_ORANGE_BTN, font=Theme.FONT_HEADING,
                     command=self.export_outlook_emails).pack(fill='x')

    def _build_actions_section(self, parent):
        SectionLabel(parent, "Actions", "🚀",
                     color=Theme.BG_GREEN_BTN).pack(fill='x', pady=(8,4))
        card = PortalCard(parent); card.pack(fill='x', pady=(0,8))
        b    = card.body
        PortalButton(b, text="Process & Upload", icon="🚀",
                     bg_color=Theme.BG_GREEN_BTN, font=Theme.FONT_HEADING,
                     command=self.process_file).pack(fill='x', pady=(0,5))
        PortalButton(b, text="Process Only (No Upload)", icon="⚙️",
                     bg_color=Theme.BG_BLUE_BTN, font=Theme.FONT_BODY,
                     command=lambda: self.process_file(upload=False)).pack(
                         fill='x', pady=(0,5))
        tip = tk.Frame(b, bg=Theme.BG_INFO_LIGHT, padx=6, pady=5)
        tip.pack(fill='x', pady=(4,0))
        tk.Label(tip,
                 text="💡 DO / GDPR / CCPA / USA / NGDPR\n"
                      "   Canada → merged into USA bucket\n"
                      "   🏢 Gartner: DO + WW (Segment split)\n"
                      "   📣 Campaign Name → Priority 1\n"
                      "   📝 Project Name  → Priority 2 fallback\n"
                      "   🔢 OLI/CID: skipped if not found\n"
                      "   ✅ & , ; etc. auto-removed from filenames\n"
                      "   📄 Output tracker Excel generated",
                 fg=Theme.FG_SECONDARY, bg=Theme.BG_INFO_LIGHT,
                 font=Theme.FONT_TINY, justify='left').pack(anchor='w')

    def _build_mapping_card(self, parent):
        card = PortalCard(parent, title="Column Mapping", icon="🔗")
        card.pack(fill='both', expand=True)
        b = card.body
        br = tk.Frame(b, bg=Theme.BG_CARD); br.pack(fill='x', pady=(0,6))
        self.auto_badge = tk.Label(br, text="  ⚡ AUTO-MAP  ",
                                   fg=Theme.FG_ACCENT, bg=Theme.BG_ACCENT_LIGHT,
                                   font=(Theme.FONT_FAMILY,7,"bold"), padx=5, pady=2)
        self.auto_badge.pack(side='left')
        self.map_info = tk.Label(br, text="Load file to auto-detect",
                                 fg=Theme.FG_MUTED, bg=Theme.BG_CARD,
                                 font=Theme.FONT_TINY)
        self.map_info.pack(side='left', padx=(8,0))
        tk.Frame(b, bg=Theme.BORDER_LIGHT, height=1).pack(fill='x', pady=(0,6))
        for field, icon, req in [
            ("Asset",                   "📄", True),
            ("Email",                   "📧", True),
            ("Country",                 "🌍", True),
            ("Campaign",                "📣", False),
            ("OLI/CID",                 "🔢", False),
            ("Segment",                 "🏷️", False),
            ("State/Region (optional)", "📍", False),
        ]:
            row = MappingRow(b, field, icon=icon, required=req)
            row.pack(fill='x')
            self.mapping_rows[field] = row

    def _build_console(self, parent):
        f = tk.Frame(parent, bg=Theme.BG_CARD,
                     highlightbackground=Theme.BORDER_CARD,
                     highlightthickness=1)
        f.pack(fill='both', expand=True)
        self.console = LogConsole(f)
        self.console.pack(fill='both', expand=True)

    # ── HELPERS ───────────────────────────────────────────────────
    def log(self, msg, tag=None):
        self.root.after(0, lambda: self.console.log(msg, tag))

    def update_stat(self, key, val):
        self.root.after(0, lambda: self.stats_cards[key].set_value(val))

    def set_status(self, text, bg, fg):
        def _u():
            self.status_frame.config(bg=bg)
            self.status_dot.config(fg=fg, bg=bg)
            self.status_lbl.config(text=text, fg=fg, bg=bg)
        self.root.after(0, _u)

    def update_login_indicator(self, connected, text=""):
        def _u():
            c = Theme.FG_SUCCESS if connected else (
                Theme.FG_DANGER if text else Theme.FG_MUTED)
            t = text or ("Connected" if connected else "Not logged in")
            self.login_dot.config(fg=c)
            self.login_lbl.config(text=t, fg=c)
            self.login_status_lbl.config(text=f"● {t}", fg=c)
        self.root.after(0, _u)

    def _validate_creds(self):
        u = self.username_var.get().strip()
        p = self.password_var.get().strip()
        if not u:
            messagebox.showwarning("Missing","Please enter Portal Email.")
            return None, None
        if not p:
            messagebox.showwarning("Missing","Please enter Portal Password.")
            return None, None
        return u, p

    def _get_ui_mapping(self):
        return {
            "email_col"  : fix_column(self.mapping_rows["Email"].get_value()),
            "asset_col"  : fix_column(self.mapping_rows["Asset"].get_value()),
            "country_col": fix_column(self.mapping_rows["Country"].get_value()),
            "camp_col"   : self.mapping_rows["Campaign"].get_value(),
            "oli_col"    : self.mapping_rows["OLI/CID"].get_value(),
            "segment_col": self.mapping_rows["Segment"].get_value(),
            "state_raw"  : self.mapping_rows["State/Region (optional)"].get_value(),
        }

    def _welcome(self):
        emp  = self.sys_user["employee_id"]
        name = self.sys_user["full_name"]
        pc   = self.sys_user["pc_name"]
        self.console.log("─"*50, "muted")
        self.console.log(
            "  SmartLead Portal  v3.4  ⚡ Portal-Safe Filenames", "header")
        self.console.log("─"*50, "muted")
        self.console.log("")
        self.console.log(f"  👤  Employee : {emp}", "success")
        if name and name != emp:
            self.console.log(f"  👤  Name     : {name}", "success")
        self.console.log(f"  💻  Computer : {pc}", "info")
        self.console.log("")
        chrome = find_chrome_path()
        if chrome:
            self.console.log("  🌐  Browser  : ✅ Found", "success")
            self.console.log(f"      {chrome}", "muted")
        else:
            self.console.log("  🌐  Browser  : ❌ NOT FOUND", "error")
            self.console.log("      https://www.google.com/chrome/", "warning")
        self.console.log("")
        self.console.log("  ✅  FILENAME SANITIZATION:", "header")
        self.console.log("  • & → removed  (was breaking portal input)", "success")
        self.console.log("  • , → removed  (was breaking portal parsing)", "success")
        self.console.log("  • ; ! @ # $ % ^ * ( ) etc. → removed", "success")
        self.console.log("  • Multiple __ → collapsed to single _", "success")
        self.console.log("")
        self.console.log("  📣  CAMPAIGN NAME RESOLUTION:", "header")
        self.console.log("  Priority 1 → 'Campaign Name' in email body", "success")
        self.console.log("  Priority 2 → 'Campaign'      in email body", "success")
        self.console.log("  Priority 3 → 'Project Name'  ← KEY FALLBACK", "project")
        self.console.log("  Priority 4 → 'Project'       ← fallback", "project")
        self.console.log("  Priority 5 → Regex sweep", "muted")
        self.console.log("")
        self.console.log("  🌍  REGION RULES:", "header")
        self.console.log("  • DO   → Germany / Switzerland / Austria", "info")
        self.console.log("  • GDPR → All other European countries", "info")
        self.console.log("  • CCPA → USA (state = California)", "info")
        self.console.log("  • USA  → USA + Canada (merged)", "info")
        self.console.log("  • NGDPR→ Everything else", "info")
        self.console.log("")

    def test_login(self):
        u, p = self._validate_creds()
        if not u: return
        self.log("━"*48, "muted")
        self.log("🔑  Testing Portal Login...", "header")
        self.set_status("Testing...", Theme.BG_WARNING_LIGHT, "#D97706")
        self.update_login_indicator(False, "Connecting...")
        def _run():
            try:
                with sync_playwright() as pw:
                    br   = self._launch_browser(pw)
                    page = br.new_page()
                    portal_login(page, u, p, log_func=self.log)
                    self.log("✅ Login test PASSED!")
                    self.update_login_indicator(True, f"Verified: {u}")
                    self.set_status("Login OK", Theme.BG_SUCCESS_LIGHT,
                                    Theme.FG_SUCCESS)
                    time.sleep(2); br.close()
            except Exception as e:
                self.log(f"❌ Login FAILED: {e}")
                self.update_login_indicator(False, "Login failed")
                self.set_status("Login Failed", Theme.BG_DANGER_LIGHT,
                                Theme.FG_DANGER)
        threading.Thread(target=_run, daemon=True).start()

    def refresh_column_menus(self, path, file_password=None):
        try:
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path)
            elif file_password:
                buf, was_enc = open_protected_excel(path, file_password)
                if was_enc and buf:
                    xl  = pd.ExcelFile(buf)
                    dfs = [xl.parse(s) for s in xl.sheet_names]
                    df  = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
                else:
                    xl  = pd.ExcelFile(path)
                    dfs = [xl.parse(s) for s in xl.sheet_names]
                    df  = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
            else:
                xl  = pd.ExcelFile(path)
                dfs = [xl.parse(s) for s in xl.sheet_names]
                df  = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
        except Exception as e:
            self.log(f"❌ Error reading file: {e}"); return

        cols   = [str(c).strip() for c in df.columns.tolist()]
        am     = auto_map_columns(cols)
        mapped = 0
        for field, row in self.mapping_rows.items():
            av = am.get(field, "")
            row.update_options(cols, av)
            if av: mapped += 1

        self.update_stat("loaded", 1)
        self.update_stat("rows", len(df))
        fname = os.path.basename(path)
        self.file_ind.config(text=f"📄 {fname}", fg=Theme.FG_PRIMARY)
        self.log(f"📂 Loaded: {fname}  ({len(df):,} rows, {len(cols)} cols)")
        if file_password:
            self.log(f"   🔐 Unlocked: {file_password}", "success")
        self.log(f"🔗 Auto-mapped {mapped}/{len(self.mapping_rows)} fields", "success")
        for field, row in self.mapping_rows.items():
            v = row.get_value()
            if v: self.log(f"   ✓ {field} → {v}", "success")
            else: self.log(f"   ○ {field} → (select manually)", "muted")

        if mapped == len(self.mapping_rows):
            self.auto_badge.config(text="  ✅ ALL MAPPED  ",
                                   bg=Theme.BG_SUCCESS_LIGHT, fg=Theme.FG_SUCCESS)
            self.map_info.config(text="All columns detected")
        elif mapped > 0:
            self.auto_badge.config(
                text=f"  ⚡ {mapped}/{len(self.mapping_rows)} MAPPED  ",
                bg=Theme.BG_WARNING_LIGHT, fg=Theme.FG_WARNING)
            self.map_info.config(text="Some need manual selection")
        else:
            self.auto_badge.config(text="  ⚠ MANUAL  ",
                                   bg=Theme.BG_DANGER_LIGHT, fg=Theme.FG_DANGER)
            self.map_info.config(text="Map columns manually")

    def browse_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV","*.xlsx *.xlsm *.csv")])
        if path:
            self.file_path_var.set(path)
            self.set_status("Loading...", Theme.BG_WARNING_LIGHT, Theme.FG_WARNING)
            self.refresh_column_menus(path)
            self.set_status("Ready", Theme.BG_SUCCESS_LIGHT, Theme.FG_SUCCESS)

    # ─────────────────────────────────────────
    # OUTLOOK SYNC
    # ─────────────────────────────────────────
    def export_outlook_emails(self):
        try:
            raw = (self.save_folder_var.get().strip() or
                   filedialog.askdirectory(title="Choose Save Folder"))
            if not raw: return
            sf = os.path.abspath(raw)
            os.makedirs(sf, exist_ok=True)
            self.save_folder_var.set(sf)
            self.log(f"📁 Base save folder: {sf}", "info")
            self.set_status("Syncing...", Theme.BG_WARNING_LIGHT, Theme.FG_WARNING)

            outlook    = win32com.client.Dispatch("Outlook.Application")
            selection  = outlook.ActiveExplorer().Selection
            mail_count = selection.Count

            if mail_count == 0:
                messagebox.showwarning("No Selection",
                    "Select at least one email in Outlook first.")
                self.set_status("Ready", Theme.BG_SUCCESS_LIGHT, Theme.FG_SUCCESS)
                return

            self.log("━"*48, "muted")
            self.log(f"📬  Found {mail_count} selected email(s)", "header")
            self.outlook_mail_jobs = []

            for mail_idx in range(1, mail_count + 1):
                try:
                    mail        = selection.Item(mail_idx)
                    raw_subject = mail.Subject or f"Email_{mail_idx}"
                    safe_subj   = safe_filename(raw_subject, max_len=60)
                    if not safe_subj:
                        safe_subj = f"Email_{mail_idx}"

                    self.log(f"\n📧 [{mail_idx}/{mail_count}] {raw_subject}", "info")

                    file_password = extract_b2bmg_password(raw_subject)
                    b2bmg_oli     = extract_b2bmg_oli(raw_subject)
                    is_b2bmg      = file_password is not None

                    if is_b2bmg:
                        self.log(f"   🔐 B2BMG — OLI: {b2bmg_oli} | pwd: {file_password}",
                                 "warning")

                    # Parse body
                    mail_body = ""
                    data      = {"campaign":"","oli":"","client":"","source":""}
                    try:
                        mail_body = mail.Body or ""
                        data      = parse_email_body(mail_body)
                    except Exception as pe:
                        self.log(f"   ⚠️  Body parse error: {pe}", "warning")

                    # Log campaign source
                    if data["campaign"]:
                        src = data.get("source","")
                        tag = "project" if "Project" in src else "success"
                        self.log(f"   📣 Campaign (via '{src}'): '{data['campaign']}'", tag)
                    else:
                        self.log("   ⚠️  Campaign Name not found in body", "warning")
                        self.log("   ⚠️  Project Name  not found in body", "warning")

                    # Resolve client
                    client_name = data.get("client","").strip()
                    if not client_name or client_name.lower() == "unknown":
                        client_name = extract_client_name(mail_body, raw_subject)

                    is_gartner = is_gartner_client(client_name)

                    # Final campaign resolution
                    campaign_name = data.get("campaign","").strip()
                    if not is_valid_campaign(campaign_name):
                        subj_clean = re.sub(
                            r'(re:|fw:|fwd:|b2bmg|\d{4})','',
                            raw_subject, flags=re.IGNORECASE).strip()
                        subj_clean = re.sub(r'\s+','_', subj_clean).strip('_')
                        campaign_name = (sanitize_for_portal(subj_clean[:80])
                                         if subj_clean else f"Campaign_{mail_idx}")
                        self.log(f"   ⚠️  Using subject as campaign: '{campaign_name}'",
                                 "warning")

                    # OLI
                    oli_val = normalize_oli(data.get("oli",""))

                    # B2BMG overrides
                    if is_b2bmg:
                        campaign_name = "B2BMG"
                        if b2bmg_oli:
                            oli_val = normalize_oli(b2bmg_oli)

                    if is_gartner:
                        self.log(f"   🏢 GARTNER — Client: '{client_name}'", "warning")

                    self.log(f"   Client  : {client_name}", "info")
                    self.log(
                        f"   Campaign: {campaign_name}  "
                        f"OLI: {oli_val if oli_val else '(none)'}", "info")

                    # Folders
                    safe_campaign = safe_folder_name(campaign_name, max_len=60)
                    if not safe_campaign:
                        safe_campaign = f"Campaign_{mail_idx}"
                    folders = create_campaign_folders(sf, safe_campaign)
                    self.log(f"   📁 Folder: {folders['campaign_folder']}", "info")
                    os.makedirs(folders["attachments_folder"], exist_ok=True)
                    os.makedirs(folders["csv_folder"],         exist_ok=True)

                    # Save .msg
                    try:
                        msg_name = safe_filename(raw_subject, max_len=60) + ".msg"
                        msg_path = os.path.normpath(
                            os.path.join(folders["attachments_folder"], msg_name))
                        mail.SaveAs(msg_path, 3)
                        self.log(f"   ✅ MSG saved: {msg_name}", "success")
                    except Exception as em:
                        self.log(f"   ⚠️  MSG save skipped: {em}", "warning")

                    # Save attachment
                    att_path = None; att_password = None
                    for att in mail.Attachments:
                        aln = att.FileName.lower()
                        if aln.endswith(('.csv','.xlsx','.xls','.xlsm')):
                            ext      = os.path.splitext(att.FileName)[1]
                            base     = os.path.splitext(att.FileName)[0]
                            safe_att = safe_filename(base, 80) + ext
                            att_path = os.path.normpath(
                                os.path.join(
                                    folders["attachments_folder"], safe_att))
                            os.makedirs(os.path.dirname(att_path), exist_ok=True)
                            try:
                                att.SaveAsFile(att_path)
                                self.log(f"   ✅ Saved: {safe_att}", "success")
                            except Exception as se:
                                self.log(f"   ❌ Save failed '{att.FileName}': {se}",
                                         "error")
                                try:
                                    fp2 = os.path.normpath(os.path.join(sf, safe_att))
                                    att.SaveAsFile(fp2)
                                    att_path = fp2
                                    self.log(f"   ✅ Fallback: {fp2}", "warning")
                                except Exception as fe:
                                    self.log(f"   ❌ Fallback failed: {fe}", "error")
                                    att_path = None

                            if att_path and os.path.exists(att_path):
                                if file_password:
                                    att_password = file_password
                                elif re.search(r'B2BMG', att.FileName, re.IGNORECASE):
                                    att_password = (
                                        extract_b2bmg_password(raw_subject) or
                                        extract_b2bmg_password(att.FileName))
                                if att_password and att_path.lower().endswith(
                                        ('.xlsx','.xlsm','.xls')):
                                    buf, was_enc = open_protected_excel(
                                        att_path, att_password)
                                    if was_enc and buf:
                                        self.log(f"   ✅ Unlocked: {att_password}",
                                                 "success")
                                    elif was_enc and not buf:
                                        self.log("   ❌ Wrong password", "error")
                                        att_password = None
                                    else:
                                        att_password = None
                            break

                    if not att_path:
                        self.log("   ⚠️  No CSV/Excel attachment found", "warning")

                    self.outlook_mail_jobs.append({
                        "index"          : mail_idx,
                        "subject"        : raw_subject,
                        "campaign"       : campaign_name,
                        "oli"            : oli_val,
                        "client_name"    : client_name,
                        "is_gartner"     : is_gartner,
                        "att_path"       : att_path,
                        "file_password"  : att_password,
                        "is_b2bmg"       : is_b2bmg,
                        "b2bmg_oli"      : b2bmg_oli,
                        "csv_folder"     : folders["csv_folder"],
                        "campaign_folder": folders["campaign_folder"],
                    })

                except Exception as me:
                    import traceback
                    self.log(f"   ❌ Email {mail_idx}: {me}", "error")
                    self.log(traceback.format_exc(), "error")

            first_job = next(
                (j for j in self.outlook_mail_jobs if j["att_path"]), None)
            if first_job and first_job["att_path"] and \
                    os.path.exists(first_job["att_path"]):
                self.file_path_var.set(first_job["att_path"])
                self.refresh_column_menus(
                    first_job["att_path"],
                    file_password=first_job.get("file_password"))

            synced = sum(1 for j in self.outlook_mail_jobs if j["att_path"])
            self.log(f"\n✅ Sync complete — "
                     f"{len(self.outlook_mail_jobs)} emails, {synced} attachments")
            self.update_stat("loaded", synced)
            self.set_status("Ready", Theme.BG_SUCCESS_LIGHT, Theme.FG_SUCCESS)

        except Exception as e:
            import traceback
            self.log(f"❌ Outlook Error: {e}", "error")
            self.log(traceback.format_exc(), "error")
            self.set_status("Error", Theme.BG_DANGER_LIGHT, Theme.FG_DANGER)

    # ─────────────────────────────────────────
    # PROCESS FILE
    # ─────────────────────────────────────────
    def process_file(self, upload=True):
        user = self.user_name_var.get().strip()
        path = self.file_path_var.get().strip()

        if not path and not self.outlook_mail_jobs:
            messagebox.showwarning("Missing File",
                "Select a Data File or sync from Outlook.")
            return

        username = password = None
        if upload:
            username, password = self._validate_creds()
            if not username: return

        sf = self.save_folder_var.get().strip()
        if not sf:
            sf = filedialog.askdirectory(title="Select Base Save Folder")
            if not sf: return
            self.save_folder_var.set(sf)
        os.makedirs(sf, exist_ok=True)
        self.set_status("Processing...", Theme.BG_WARNING_LIGHT, Theme.FG_WARNING)

        self.tracking_records = []

        if self.outlook_mail_jobs:
            jobs = list(self.outlook_mail_jobs)
        else:
            camp    = extracted_metadata.get("campaign","") or ""
            oli_raw = extracted_metadata.get("oli","")    or ""
            if not is_valid_campaign(camp):
                camp = "Unknown_Campaign"
            oli_val = normalize_oli(oli_raw)
            folders = create_campaign_folders(sf, camp)
            jobs = [{
                "index"          : 1,
                "subject"        : os.path.basename(path),
                "campaign"       : camp,
                "oli"            : oli_val,
                "client_name"    : "Unknown",
                "is_gartner"     : False,
                "att_path"       : path,
                "file_password"  : None,
                "is_b2bmg"       : False,
                "b2bmg_oli"      : None,
                "csv_folder"     : folders["csv_folder"],
                "campaign_folder": folders["campaign_folder"],
            }]

        ui_map = self._get_ui_mapping()
        self.log("━"*48, "muted")
        self.log(f"🚀  Processing {len(jobs)} job(s)  |  User: {user}", "header")
        self.log("━"*48, "muted")

        all_csv_tasks = []
        total_rows    = 0
        jobs_ok       = 0
        jobs_fail     = 0
        emp_id        = self.sys_user["employee_id"]
        process_time  = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        # ── GARTNER PROCESSOR ──────────────────────────────────────
        def _process_gartner_job(job_idx, job, df,
                                  email_col, asset_col, country_col,
                                  segment_col, csv_folder,
                                  subject, client_name, campaign_from_email):
            job_tasks     = []
            mail_campaign = str(campaign_from_email).strip()
            if not is_valid_campaign(mail_campaign):
                mail_campaign = f"GartnerCampaign_{job_idx}"

            self.log("   🏢 GARTNER MODE", "warning")
            self.log(f"   📣 Campaign: '{mail_campaign}'", "info")

            resolved_seg_col = segment_col
            if not resolved_seg_col or resolved_seg_col not in df.columns:
                jc  = [str(c).strip() for c in df.columns.tolist()]
                jam = auto_map_columns(jc)
                resolved_seg_col = jam.get("Segment","")
            if not resolved_seg_col or resolved_seg_col not in df.columns:
                for col in df.columns:
                    if str(col).strip().lower() in [
                        "campaignname","campaign_name","campaign name",
                        "segment","segment_name","segmentname",
                    ]:
                        resolved_seg_col = str(col).strip(); break

            if not resolved_seg_col or resolved_seg_col not in df.columns:
                self.log("   ❌ Segment column not found — skipping.", "error")
                return job_tasks

            self.log(f"   🏷️  Segment col: '{resolved_seg_col}'", "info")

            cs    = df[country_col].astype(str).str.lower().str.strip()
            is_do = cs.isin(DO_COUNTRIES)

            for asset in df[asset_col].dropna().unique():
                adf        = df[df[asset_col] == asset].copy()
                a_is_do    = is_do[adf.index]
                full_asset = str(asset).strip()
                segments   = adf[resolved_seg_col].dropna().unique()
                if len(segments) == 0:
                    segments = ["Unknown_Segment"]

                for segment in segments:
                    seg_full  = str(segment).strip()
                    seg_short = get_segment_short_name(seg_full)
                    self.log(f"   🏷️  Seg: '{seg_short}'", "muted")

                    seg_df    = (adf.copy() if seg_full == "Unknown_Segment"
                                 else adf[adf[resolved_seg_col].astype(str)
                                          == seg_full].copy())
                    seg_is_do = a_is_do[seg_df.index]

                    do_df = seg_df[seg_is_do]
                    if not do_df.empty:
                        out_do              = pd.DataFrame({'Email': do_df[email_col].unique()})
                        out_do['Asset']     = full_asset
                        out_do['Segment']   = seg_full
                        out_do['FirstName'] = ''
                        out_do['LastName']  = ''
                        out_do['City']      = ''
                        out_do['Country']   = ''
                        fname_do = format_filename_gartner(
                            mail_campaign, seg_full, full_asset, "DO", len(out_do))
                        fpath_do = os.path.join(csv_folder, fname_do)
                        out_do.to_csv(fpath_do, index=False)
                        job_tasks.append((fpath_do, fname_do, subject))
                        self.log(f"   💾 [DO] {fname_do}  ({len(out_do)} rows)")
                        self.tracking_records.append({
                            "EMP ID"        : emp_id,
                            "Process Date"  : process_time,
                            "Email Subject" : subject,
                            "Client Name"   : client_name,
                            "Campaign Name" : mail_campaign,
                            "OLI/CID"       : "",
                            "Segment"       : seg_full,
                            "Asset Name"    : full_asset,
                            "Region"        : "DO",
                            "CSV File Name" : fname_do,
                            "Data Count"    : len(out_do),
                            "Source File"   : os.path.basename(job.get("att_path","")),
                            "B2BMG"         : "No",
                            "Uploaded"      : "Pending" if upload else "No Upload",
                            "CSV Path"      : fpath_do,
                        })

                    ww_df = seg_df[~seg_is_do]
                    if not ww_df.empty:
                        out_ww              = pd.DataFrame({'Email': ww_df[email_col].unique()})
                        out_ww['Asset']     = full_asset
                        out_ww['Segment']   = seg_full
                        out_ww['FirstName'] = ''
                        out_ww['LastName']  = ''
                        out_ww['City']      = ''
                        out_ww['Country']   = ''
                        fname_ww = format_filename_gartner(
                            mail_campaign, seg_full, full_asset, "WW", len(out_ww))
                        fpath_ww = os.path.join(csv_folder, fname_ww)
                        out_ww.to_csv(fpath_ww, index=False)
                        job_tasks.append((fpath_ww, fname_ww, subject))
                        self.log(f"   💾 [WW] {fname_ww}  ({len(out_ww)} rows)")
                        self.tracking_records.append({
                            "EMP ID"        : emp_id,
                            "Process Date"  : process_time,
                            "Email Subject" : subject,
                            "Client Name"   : client_name,
                            "Campaign Name" : mail_campaign,
                            "OLI/CID"       : "",
                            "Segment"       : seg_full,
                            "Asset Name"    : full_asset,
                            "Region"        : "WW",
                            "CSV File Name" : fname_ww,
                            "Data Count"    : len(out_ww),
                            "Source File"   : os.path.basename(job.get("att_path","")),
                            "B2BMG"         : "No",
                            "Uploaded"      : "Pending" if upload else "No Upload",
                            "CSV Path"      : fpath_ww,
                        })

            return job_tasks

        # ── STANDARD PROCESSOR ─────────────────────────────────────
        def _process_standard_job(job_idx, job, df,
                                   email_col, asset_col, country_col,
                                   camp_col, oli_col, state_col,
                                   campaign, oli, csv_folder,
                                   subject, client_name, is_b2bmg, b2bmg_oli):
            job_tasks = []

            cs = df[country_col].astype(str).str.lower().str.strip()
            ss = (df[state_col].astype(str).str.lower().str.strip()
                  if state_col and state_col in df.columns
                  else pd.Series("", index=df.index))

            is_do     = cs.isin(DO_COUNTRIES)
            is_euro   = cs.isin(EUROPE_COUNTRIES)
            is_us     = cs.isin(US_VARIANTS)
            is_canada = cs.isin(CANADA_VARIANTS)
            is_ca_st  = is_us & ss.str.contains(
                r'\b(ca|california)\b', regex=True, na=False)

            region = pd.Series("NGDPR", index=df.index)
            region[is_canada]         = "USA_CANADA"
            region[is_us & ~is_ca_st] = "USA_CANADA"
            region[is_us & is_ca_st]  = "CCPA"
            region[is_euro & ~is_do]  = "GDPR"
            region[is_do]             = "DO"

            has_gdpr = (region == "GDPR").any()
            if has_gdpr:
                region[region == "USA_CANADA"] = "USA"
            else:
                region[region == "USA_CANADA"] = "NGDPR"

            for asset in df[asset_col].dropna().unique():
                adf = df[df[asset_col] == asset].copy()
                ar  = region[adf.index]

                full_asset          = str(asset).strip()
                asset_name_for_file = get_asset_name_no_stopwords(
                    full_asset, max_words=5)

                self.log(f"   📝 Asset: '{asset_name_for_file}'", "muted")
                self.log(
                    f"   📊 DO:{(ar=='DO').sum()}  GDPR:{(ar=='GDPR').sum()}  "
                    f"CCPA:{(ar=='CCPA').sum()}  USA:{(ar=='USA').sum()}  "
                    f"NGDPR:{(ar=='NGDPR').sum()}", "muted")

                active_tags = [t for t in ["DO","GDPR","CCPA","USA","NGDPR"]
                               if (ar == t).any()]

                for tag in active_tags:
                    sdf = adf[ar == tag]
                    if sdf.empty: continue

                    out              = pd.DataFrame({'Email': sdf[email_col].unique()})
                    out['Asset']     = full_asset
                    out['FirstName'] = ''
                    out['LastName']  = ''
                    out['City']      = ''
                    out['Country']   = ''

                    # Campaign
                    if camp_col and camp_col in sdf.columns:
                        cv = str(sdf[camp_col].iloc[0]).strip()
                        if not is_valid_campaign(cv):
                            cv = campaign
                    else:
                        cv = campaign
                    if not is_valid_campaign(cv):
                        cv = f"Campaign_{job_idx}"

                    # OLI — never 0000
                    if oli_col and oli_col in sdf.columns:
                        ov = normalize_oli(sdf[oli_col].iloc[0])
                    else:
                        ov = normalize_oli(oli)
                    if is_b2bmg and b2bmg_oli:
                        ov = normalize_oli(b2bmg_oli)

                    fname = format_filename(
                        cv, ov, asset_name_for_file, tag, len(out),
                        is_b2bmg=is_b2bmg, b2bmg_oli=b2bmg_oli)
                    fpath = os.path.join(csv_folder, fname)
                    out.to_csv(fpath, index=False)
                    job_tasks.append((fpath, fname, subject))

                    self.log(
                        f"   💾 [{tag}] {fname}  ({len(out)} rows)  "
                        f"OLI: {ov if ov else '(none)'}")

                    self.tracking_records.append({
                        "EMP ID"       : emp_id,
                        "Process Date" : process_time,
                        "Email Subject": subject,
                        "Client Name"  : client_name,
                        "Campaign Name": (cv if not is_b2bmg else "B2BMG"),
                        "OLI/CID"      : (ov if not is_b2bmg
                                          else (normalize_oli(b2bmg_oli) or ov)),
                        "Segment"      : "",
                        "Asset Name"   : full_asset,
                        "Region"       : tag,
                        "CSV File Name": fname,
                        "Data Count"   : len(out),
                        "Source File"  : (os.path.basename(job.get("att_path",""))
                                          if job.get("att_path") else ""),
                        "B2BMG"        : "Yes" if is_b2bmg else "No",
                        "Uploaded"     : "Pending" if upload else "No Upload",
                        "CSV Path"     : fpath,
                    })

            return job_tasks

        # ── MAIN JOB LOOP ──────────────────────────────────────────
        def _process_one_job(job_idx, job):
            nonlocal total_rows, jobs_ok, jobs_fail

            att_path      = job.get("att_path")
            campaign      = job.get("campaign","")
            oli           = normalize_oli(job.get("oli",""))
            subject       = job.get("subject","")
            csv_folder    = job.get("csv_folder", sf)
            file_password = job.get("file_password")
            is_b2bmg      = job.get("is_b2bmg", False)
            b2bmg_oli     = job.get("b2bmg_oli")
            client_name   = job.get("client_name","Unknown")
            is_gartner    = job.get("is_gartner", False)

            if not is_valid_campaign(campaign):
                campaign = f"Campaign_{job_idx}"

            self.log(f"\n{'─'*44}", "muted")
            self.log(f"📧 Job {job_idx}/{len(jobs)}: {subject}", "header")
            self.log(f"   📣 Campaign : {campaign}", "info")
            self.log(f"   🔢 OLI/CID  : {oli if oli else '(none)'}", "info")

            if not att_path:
                self.log("   ⚠️  No attachment — skipping", "warning")
                return []
            if not os.path.exists(att_path):
                self.log(f"   ❌ File not found: {att_path}", "error")
                return []

            if not is_gartner:
                is_gartner = is_gartner_client(client_name)

            job_cols = get_columns_from_file(att_path, file_password)
            job_am   = auto_map_columns(job_cols) if job_cols else {}

            def _resolve(field, ui_val):
                ui_val   = (ui_val or "").strip()
                auto_val = job_am.get(field,"").strip()
                if ui_val and ui_val in job_cols:
                    self.log(f"      🔧 {field}: manual → '{ui_val}'", "manual")
                    return ui_val
                if auto_val:
                    self.log(f"      ⚡ {field}: auto   → '{auto_val}'", "muted")
                    return auto_val
                return ""

            email_col   = fix_column(_resolve("Email",   ui_map["email_col"]))
            asset_col   = fix_column(_resolve("Asset",   ui_map["asset_col"]))
            country_col = fix_column(_resolve("Country", ui_map["country_col"]))
            camp_col    = _resolve("Campaign",                ui_map["camp_col"])
            oli_col     = _resolve("OLI/CID",                ui_map["oli_col"])
            segment_col = _resolve("Segment",                 ui_map["segment_col"])
            state_raw   = _resolve("State/Region (optional)", ui_map["state_raw"])
            state_col   = fix_column(state_raw) if state_raw.strip() else None

            missing = [f for f,c in [
                ("Email",  email_col),
                ("Asset",  asset_col),
                ("Country",country_col),
            ] if not c]
            if missing:
                self.log(f"   ❌ Missing columns: {missing} — skipping", "error")
                return []

            df = get_highlighted_data(att_path, email_col,
                                      file_password=file_password,
                                      log_func=self.log)
            if df is None or df.empty:
                self.log("   ⚠️  No highlighted rows — skipping", "warning")
                return []

            for cn, cv in [("Email",email_col),("Asset",asset_col),
                           ("Country",country_col)]:
                if cv not in df.columns:
                    self.log(f"   ❌ Column '{cv}' not in data — skipping","error")
                    return []

            total_rows += len(df)
            self.log(f"   📊 {len(df)} highlighted rows")

            if is_gartner:
                return _process_gartner_job(
                    job_idx, job, df,
                    email_col, asset_col, country_col,
                    segment_col, csv_folder,
                    subject, client_name, campaign)
            else:
                return _process_standard_job(
                    job_idx, job, df,
                    email_col, asset_col, country_col,
                    camp_col, oli_col, state_col,
                    campaign, oli, csv_folder,
                    subject, client_name, is_b2bmg, b2bmg_oli)

        for job_idx, job in enumerate(jobs, 1):
            try:
                tasks = _process_one_job(job_idx, job)
                all_csv_tasks.extend(tasks)
                if tasks: jobs_ok  += 1
                else:     jobs_fail += 1
            except Exception as ex:
                jobs_fail += 1
                import traceback
                self.log(f"❌ Job {job_idx} crashed: {ex}", "error")
                self.log(traceback.format_exc(), "error")

        self.update_stat("rows", total_rows)
        self.log("")
        self.log("━"*48, "muted")
        self.log(f"📊  Jobs: ✅ {jobs_ok}  ❌ {jobs_fail}  "
                 f"|  CSVs: {len(all_csv_tasks)}", "header")
        self.log("━"*48, "muted")

        if upload and all_csv_tasks:
            self.log(f"\n🚀 Starting upload ({len(all_csv_tasks)} files)...")
            threading.Thread(
                target=self._run_automation,
                args=(sf, username, password, all_csv_tasks),
                daemon=True).start()
        else:
            self._save_output_tracker(sf)
            self.set_status(
                "Done" if jobs_fail == 0 else f"{jobs_fail} jobs failed",
                Theme.BG_SUCCESS_LIGHT if jobs_fail == 0 else Theme.BG_WARNING_LIGHT,
                Theme.FG_SUCCESS if jobs_fail == 0 else Theme.FG_WARNING)
            self.log("✅ Done — files saved (no upload)")

    def _save_output_tracker(self, save_folder, upload_results=None):
        if not self.tracking_records:
            self.log("   ⚠️  No tracking records to save", "warning")
            return None

        if upload_results:
            for rec in self.tracking_records:
                fname = rec.get("CSV File Name","")
                if fname in upload_results:
                    rec["Uploaded"] = upload_results[fname]

        df = pd.DataFrame(self.tracking_records)
        col_order = [
            "EMP ID","Process Date","Email Subject","Client Name",
            "Campaign Name","OLI/CID","Segment","Asset Name","Region",
            "CSV File Name","Data Count","Source File",
            "B2BMG","Uploaded","CSV Path",
        ]
        col_order = [c for c in col_order if c in df.columns]
        df = df[col_order]

        d            = datetime.now().strftime("%d_%m_%y_%H%M")
        emp          = self.sys_user["employee_id"]
        tracker_name = f"Output_Tracker_{emp}_{d}.xlsx"
        tracker_path = os.path.join(save_folder, tracker_name)
        df.to_excel(tracker_path, index=False, engine='openpyxl')

        self.log(f"\n📄 Output Tracker saved: {tracker_name}", "success")
        self.log(f"   📁 Path: {tracker_path}", "info")
        self.log(f"   📊 Total records: {len(df)}", "info")
        self.log(f"   📋 Total data count: {df['Data Count'].sum():,}", "info")
        return tracker_path

    def _run_automation(self, save_folder, username, password,
                        all_csv_tasks=None):
        self.set_status("Uploading...", Theme.BG_WARNING_LIGHT, Theme.FG_WARNING)
        self.log("━"*48, "muted")
        self.log("🚀  Upload Automation", "header")
        self.log("━"*48, "muted")

        files_to_upload = all_csv_tasks or [
            (os.path.join(save_folder,f), f, "")
            for f in os.listdir(save_folder)
            if (f.endswith(".csv") and
                not f.lower().startswith("summary") and
                not f.lower().startswith("output_tracker"))
        ]

        if not files_to_upload:
            self.log("⚠️  No CSV files to upload.", "warning")
            self.set_status("No files", Theme.BG_WARNING_LIGHT, Theme.FG_WARNING)
            return

        total_files    = len(files_to_upload)
        success        = 0
        fail           = 0
        upload_results = {}

        try:
            with sync_playwright() as pw:
                browser = self._launch_browser(pw)
                page    = browser.new_page()
                self.log("🔐 Logging in...", "info")
                portal_login(page, username, password, log_func=self.log)
                self.update_login_indicator(True, f"Logged in: {username}")
                self.log(f"✅ Logged in — {total_files} file(s) to upload\n")

                current_subj = ""
                for i,(fp,file,mail_subj) in enumerate(files_to_upload, 1):
                    try:
                        name = os.path.splitext(file)[0]
                        if mail_subj and mail_subj != current_subj:
                            current_subj = mail_subj
                            self.log(f"\n{'═'*40}", "muted")
                            self.log(f"📧 {mail_subj}", "header")
                            self.log(f"{'═'*40}", "muted")

                        self.log(f"\n[{i}/{total_files}] {name}")
                        page.goto(PORTAL_CONFIG["lists_url"])
                        page.wait_for_load_state("networkidle", timeout=15000)
                        page.wait_for_timeout(int(WAIT_AFTER_NAV*1000))

                        click_create_new_list(page, log_func=self.log)
                        page.wait_for_timeout(int(WAIT_AFTER_CREATE_CLICK*1000))

                        click_single_list_from_dropdown(page, log_func=self.log)
                        page.wait_for_load_state("networkidle", timeout=10000)
                        page.wait_for_timeout(int(WAIT_AFTER_SINGLE_LIST*1000))

                        # name is already portal-safe (sanitize_for_portal ran during generation)
                        fill_list_name(page, name, log_func=self.log)

                        try:
                            ta = page.locator("textarea").first
                            if ta.is_visible(timeout=1500):
                                ta.click(); ta.fill("NA")
                        except: pass

                        for sel in [
                            'button:has-text("CREATE")',
                            'button:has-text("Create")',
                            'button:has-text("SAVE")',
                            'button:has-text("Save")',
                            'button:has-text("Next")',
                            'button[type="submit"]',
                            'input[value="CREATE"]',
                            'input[value="Create"]',
                        ]:
                            try:
                                btn = page.locator(sel).first
                                if btn.is_visible(timeout=2000):
                                    btn.scroll_into_view_if_needed()
                                    btn.click()
                                    self.log("   ✅ CREATE clicked")
                                    break
                            except: continue

                        page.wait_for_load_state("networkidle", timeout=8000)
                        page.wait_for_timeout(int(WAIT_AFTER_CREATE_BTN*1000))

                        try:
                            fi = page.locator('input[type="file"]').first
                            if fi.is_visible(timeout=4000):
                                fi.set_input_files(fp)
                            else:
                                page.set_input_files('input[type="file"]', fp)
                            self.log(f"   ✅ File attached: {file}")
                        except Exception as fe:
                            self.log(f"   ⚠️  Attach: {fe}", "warning")

                        page.wait_for_timeout(int(WAIT_AFTER_FILE_ATTACH*1000))

                        import_clicked = False
                        for sel in [
                            "input[value='Import Subscribers']",
                            "input[value='Import subscribers']",
                            "button:has-text('Import Subscribers')",
                            "button:has-text('Import')",
                            "button:has-text('Upload')",
                            "input[value='Upload']",
                        ]:
                            try:
                                btn = page.locator(sel).first
                                if btn.is_visible(timeout=2500):
                                    btn.scroll_into_view_if_needed()
                                    btn.click(force=True)
                                    self.log("   ✅ Import clicked")
                                    import_clicked = True; break
                            except: continue
                        if not import_clicked:
                            self.log("   ⚠️  Import button not found", "warning")

                        page.wait_for_timeout(int(WAIT_AFTER_IMPORT*1000))
                        success += 1
                        self.update_stat("success", success)
                        self.log(f"✅ [{i}/{total_files}] {name}")
                        upload_results[file] = "Success"

                    except Exception as e:
                        fail += 1
                        self.update_stat("failed", fail)
                        self.log(f"❌ [{i}/{total_files}] {e}")
                        upload_results[file] = "Failed"
                        try:
                            shot = os.path.join(
                                save_folder,
                                f"err_{i}_{os.path.splitext(file)[0][:25]}.png")
                            page.screenshot(path=shot, full_page=True)
                            self.log(f"   📸 → {shot}", "warning")
                        except: pass

                browser.close()

            self.log("")
            self.log("━"*48, "muted")
            self.log(f"📊  ✅ {success}  |  ❌ {fail}", "header")
            self.log("━"*48, "muted")

            self._save_output_tracker(save_folder, upload_results=upload_results)

            self.set_status(
                "Complete ✓" if fail == 0 else f"{fail} Failed",
                Theme.BG_SUCCESS_LIGHT if fail == 0 else Theme.BG_DANGER_LIGHT,
                Theme.FG_SUCCESS if fail == 0 else Theme.FG_DANGER)

        except Exception as e:
            import traceback
            self.log(f"❌ Critical: {e}")
            self.log(traceback.format_exc(), "error")
            self._save_output_tracker(save_folder, upload_results=upload_results)
            self.set_status("Error", Theme.BG_DANGER_LIGHT, Theme.FG_DANGER)
            self.update_login_indicator(False, "Error")

# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = SmartLeadApp(root)
    root.mainloop()
#-------------------------update-------
import os

# ============================================================
# SAFE CSV WRITE (Prevents FileNotFoundError on user PC)
# ============================================================

try:
    output_dir = os.path.dirname(output_path)

    # 1️⃣ Check if main network path is accessible
    if not os.path.exists(r'\\DBSLFLSVR\InternalResearch'):
        raise Exception(
            "Network drive not accessible.\n"
            "Please connect to company network/VPN.\n"
            "Path: \\\\DBSLFLSVR\\InternalResearch"
        )

    # 2️⃣ Create missing folders automatically
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # 3️⃣ Save CSV safely
    safe_to_csv(df, output_path, index=False, encoding='utf-8-sig')

except PermissionError:
    raise Exception(
        f"No permission to create/write folder:\n{output_dir}\n"
        "Please contact IT team."
    )

except Exception as e:
    raise Exception(f"Error while saving CSV file:\n{str(e)}")
import os
from pathlib import Path

def safe_to_csv(df, file_path, **kwargs):
    """
    Safely writes CSV:
    - Creates directory if missing
    - Handles long Windows/UNC paths
    """
    # Create directory if it doesn't exist
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)

    # Handle long Windows paths
    if os.name == 'nt':
        if file_path.startswith('\\\\'):
            # Convert UNC path
            file_path = '\\\\?\\UNC\\' + file_path[2:]
        elif not file_path.startswith('\\\\?\\'):
            file_path = '\\\\?\\' + os.path.abspath(file_path)

    safe_to_csv(file_path, **kwargs)
    return file_path

